""" C3DC ETL File Creator """
import csv
import hashlib
import io
import json
import logging
import logging.config
import os
import pathlib
import random
import re
import sys
import uuid
import warnings

import dotenv
import jsonschema
from jsonschema import ValidationError
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from c3dc_etl_model_node import C3dcEtlModelNode
from c3dc_row_mapped_builder import C3dcRowMappedBuilder


def look_up_and_append_sys_path(*args: tuple[str, ...]) -> None:
    """ Append specified dir_name to sys path for import """
    dir_to_find: str
    for dir_to_find in args:
        parent: pathlib.Path
        for parent in pathlib.Path(os.getcwd()).parents:
            peer_dirs: list[os.DirEntry] = [d for d in os.scandir(parent) if d.is_dir()]
            path_to_append: str = next((p.path for p in peer_dirs if p.name == dir_to_find), None)
            if path_to_append:
                if path_to_append not in sys.path:
                    sys.path.append(path_to_append)
                break
look_up_and_append_sys_path('file_manager')
from c3dc_file_manager import C3dcFileManager # pylint: disable=wrong-import-position; # type: ignore


# suppress openpyxl warning about inability to parse header/footer
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

_logger = logging.getLogger(__name__)
if _logger.hasHandlers():
    _logger.handlers.clear()

logging.config.dictConfig({
    "version": 1,
    "disable_existing_loggers": True,
    "formatters": {
        "standard": {
            "format": "%(asctime)s [%(levelname)s]: %(message)s"
        }
    },
    "handlers": {
        "console": {
            "level": "DEBUG",
            "formatter": "standard",
            "class": "logging.StreamHandler",
            "stream": "ext://sys.stdout",  # Default is stderr
        },
        "file": {
            "level": "DEBUG",
            "formatter": "standard",
            "class": "logging.FileHandler",
            "filename": "c3dc_etl.log",
            "mode": "w"
        }
    },
    "loggers": {
        "": { # root logger
            "handlers": ["console", "file"],
            "level": "DEBUG",
            "propagate": False
        },
        "__main__": {
            "handlers": ["console", "file"],
            "level": "DEBUG",
            "propagate": False
        }
    }
})


class C3dcEtl:
    """ Build C3DC json ETL file from (XLSX) source file """
    TYPE_NAME_CLASS_MAP: dict[str, type | list[type]] = {
        'array': list,
        'integer': int,
        'string': str
    }
    MULTIPLE_VALUE_DELIMITER: str = ';'
    ETHNICITY_ALLOWED_VALUES: set[str] = {'Hispanic or Latino'}
    RACE_UNDETERMINED_VALUES: set[str] = {'Not Allowed to Collect', 'Not Reported', 'Unknown'}

    def __init__(self, config: dict[str, str]) -> None:
        self._config: dict[str, str] = config
        self._verify_config()

        self._json_schema_url: str = config.get('JSON_SCHEMA_URL')
        self._json_schema: dict[str, any] = {}
        self._json_schema_nodes: dict[str, any] = {}
        self._json_schema_property_enum_values: dict[str, list[str]] = {}
        self._json_schema_property_enum_code_values: dict[str, dict[str, str]] = {}
        self._json_etl_data_sets: dict[str, any] = {}
        self._raw_etl_data_objects: dict[str, dict[str, list[dict[str, any]]]] = {}
        self._random: random.Random = random.Random()
        self._c3dc_file_manager: C3dcFileManager = C3dcFileManager()

        # These are the schema properties for which 'sub' source records may be needed as they're enum/string
        # properties for which source values may be delimited (';'). Enum/string properties in the schema where
        # the allowed values may contain ';' (diagnosis.diagnosis) will be excluded
        self._sub_source_record_enum_properties: dict[str, str] = {}

        # Remote study config should contain env-agnostic info like mappings, local study config
        # should contain info specific to the local env like file paths; remote and local will be
        # merged together to form the final study configuration object
        self._study_configurations: list[dict[str, any]] = json.loads(config.get('STUDY_CONFIGURATIONS', '[]'))
        self._study_configurations = [sc for sc in self._study_configurations if sc.get('active', True)]

        # track mappings for row-mapped nodes (ex: treatments and treatment responses) separately for each
        # transformation as nested dicts: {node type => {transformation name => list of mappings}}
        self._row_mapped_node_mappings: dict[str, dict[str, list[dict[str, any]]]] = {}
        self._row_mapped_node_builders: dict[str, any] = {}
        row_mapped_node: C3dcEtlModelNode
        for row_mapped_node in C3dcRowMappedBuilder.NODE_SOURCE_VARIABLE_FIELDS:
            builder: C3dcRowMappedBuilder = C3dcRowMappedBuilder.get_instance(row_mapped_node)
            builder.generate_uuid_callback = self._generate_uuid
            builder.convert_output_value_callback = self._get_json_schema_node_property_converted_value
            builder.is_output_property_required_callback = self._is_json_schema_node_property_required
            builder.logger = _logger

            self._row_mapped_node_builders[row_mapped_node] = builder
            self._row_mapped_node_mappings[row_mapped_node] = {}

        self._source_field_output_types: dict[str, str] = {}

        self.load_json_schema()
        self.load_transformations()
        self._assert_valid_study_configurations()

    @property
    def json_schema(self) -> dict[str, any]:
        """ Get internal JSON schema, loading if needed """
        return self._json_schema if self._json_schema else self.load_json_schema()

    @property
    def json_etl_data(self) -> dict[str, any]:
        """ Get internal JSON ETL data object, building if needed """
        if not self._json_etl_data_sets:
            self.create_json_etl_files()
        return self._json_etl_data_sets

    @property
    def raw_etl_data_objects(self) -> dict[str, dict[str, list[dict[str, any]]]]:
        """ Get internal source data objects, loading if needed """
        if not self._raw_etl_data_objects:
            study_configuration: dict[str, any]
            for study_configuration in self._study_configurations:
                self._load_study_data(study_configuration)
        return self._raw_etl_data_objects

    @staticmethod
    def is_number(value: str) -> bool:
        """ Determine whether specified string is number (float or int) """
        try:
            float(value)
        except (TypeError, ValueError):
            return False
        return True

    @staticmethod
    def is_allowed_value(value: any, allowed_values: set[any]) -> bool:
        """ Determine whether specified value is contained in or subset of specified allowed values """
        return (
            allowed_values
            and
            (
                (not isinstance(value, (list, set, tuple)) and value in allowed_values)
                or
                (isinstance(value, (list, set, tuple)) and value and set(value or {}).issubset(allowed_values))
            )
        )

    @staticmethod
    def is_replacement_match(
        source_field: str,
        source_record: dict[str, any],
        old_value: str
    ) -> bool:
        """ Determine whether the source value should be replaced for the specified replacement entry """
        # source field is single field
        if not (source_field.startswith('[') and source_field.endswith(']')):
            old_value = '' if old_value is None else str(old_value).strip().casefold()
            source_value: str = source_record.get(source_field, '')
            source_value = '' if source_value is None else str(source_value).strip().casefold()
            return (
                source_field == '[string_literal]'
                or
                old_value == '*'
                or
                (old_value == '+' and source_value != '')
                or
                source_value and old_value and source_value == old_value
            )

        # source field contains multiple fields, e.g. [race, ethnicity], so check to
        # see if old values to be replaced match source values by ordinal position
        source_field_names: list[str] = [s.strip() for s in next(csv.reader([source_field.strip(' []')]))]
        old_values: list[str] = (
            list(next(csv.reader([old_value.strip(' []')], delimiter=C3dcEtl.MULTIPLE_VALUE_DELIMITER)))
        ) if old_value not in ('*', '+') else [old_value for n in source_field_names]
        if len(old_values) != len(source_field_names):
            raise RuntimeError(
                'Invalid replacement entry, number of old values to be replaced must match number of (compound) ' +
                f'source fields: old value => "{old_value}", source field => "{source_field}"'
            )
        index: int
        source_field_name: str
        for index, source_field_name in enumerate(source_field_names):
            old_val: str = old_values[index]
            old_val = '' if old_val is None else str(old_val).strip().casefold()
            src_val: str = source_record.get(source_field_name, '')
            src_val = '' if src_val is None else str(src_val).strip().casefold()

            if not (old_val == '*' or (old_val == '+' and src_val != '') or src_val == old_val):
                return False
        return True

    @staticmethod
    def collate_form_data(ordered_pairs: list[tuple[any, any]]) -> any:
        """
        Callback for object_pairs_hook arg of json.load and json.loads. Collate form 'data' elements into
        lists to avoid default behavior that only retains last element value when duplicates present
        """
        obj: dict[str, any] = {}
        key: str
        val: any
        for key, val in ordered_pairs:
            if key == 'data' and isinstance(val, list) and any(k == 'form_id' for k, _ in ordered_pairs):
                # store 'data' element as list of lists, usually single-element except when dupes present
                obj[key] = obj.get(key, [])
                obj[key].append(val)
            else:
                obj[key] = val
        return obj

    def load_transformations(self, save_local_copy: bool = False) -> list[dict[str, any]]:
        """ Download JSON transformations from configured URLs and merge with local config """
        # enumerate each per-study transformation config object in local config
        st_index: int
        study_config: list[dict[str, any]]
        for st_index, study_config in enumerate(self._study_configurations):
            # load remote study config containing transformations and mappings
            remote_transforms: dict[str, any] = json.loads(
                self._c3dc_file_manager.read_file(study_config.get('transformations_url')).decode('utf-8')
            )
            if save_local_copy:
                with open(
                    f'./{self._c3dc_file_manager.get_basename(study_config.get("transformations_url"))}',
                    'wb'
                ) as file:
                    json.dump(remote_transforms, file)

            # match by 'name' to merge remote with local transformation config sections containing e.g. paths
            remote_transforms = [t for t in remote_transforms.get('transformations') if t.get('active', True)]
            rt_index: int
            remote_transform: dict[str, any]
            for rt_index, remote_transform in enumerate(remote_transforms):
                transform_config: dict[str, any] = next(
                    (
                        t for t in study_config.get('transformations', [])
                            if t.get('name') and t.get('name') == remote_transform.get('name')
                    ),
                    {}
                )
                if not transform_config:
                    raise RuntimeError(f'No local match for remote transformation "{remote_transform.get("name")}"')

                _logger.info('Updating transformation at index %d for study configuration %d', rt_index, st_index)
                transform_config.update(remote_transform)

                # populate internal cache of source field=>output type pairs (non-compound mappings only)
                mapping: dict[str, any]
                for mapping in remote_transform.get('mappings', []):
                    source_field: str = mapping.get('source_field')
                    if source_field == '[string_literal]':
                        continue
                    if source_field.startswith('[') and source_field.endswith(']'):
                        # composite/derived source field, strip extra spaces; note csv
                        # module parses "field1, field2" into ["field1", " field2"]
                        sub_fields: list[str] = [s.strip() for s in next(
                            csv.reader([source_field.strip(' []')]))
                        ]
                        sub_field: str
                        for sub_field in sub_fields:
                            self._source_field_output_types[sub_field] = (
                                self._get_json_schema_node_property_type(mapping.get('output_field'))
                            )
                    else:
                        self._source_field_output_types[source_field] = (
                            self._get_json_schema_node_property_type(mapping.get('output_field'))
                        )

                # load row-mapped node mappings if specified in transform config
                row_mapped_node: C3dcEtlModelNode
                row_mapped_builder: C3dcRowMappedBuilder
                for row_mapped_node, row_mapped_builder in self._row_mapped_node_builders.items():
                    if f'{row_mapped_node}_mappings_path' not in transform_config:
                        _logger.info('No "%s" mappings found, skipping row-mapped mapping ETL', row_mapped_node)
                        continue

                    row_mapped_mappings: list[dict[str, any]] = self._get_row_mapped_node_mappings(
                        transform_config,
                        row_mapped_node
                    )
                    self._row_mapped_node_mappings[row_mapped_node][transform_config['name']] = row_mapped_mappings
                    row_mapped_builder.mappings = row_mapped_mappings

                    # row-mapped mappings like treatment and treatment response aren't 1:1 mappings of source to
                    # output fields however we still need to add the mapping source fields to
                    # self._source_field_output_types to make sure that the mapped source field values are loaded
                    # by self._load_source_data for harmonization
                    row_mapped_source_fields: list[str] = row_mapped_builder.get_mapped_source_fields()
                    row_mapped_source_field: str
                    for row_mapped_source_field in row_mapped_source_fields:
                        self._source_field_output_types[row_mapped_source_field] = 'string'

            # log warnings for any unmatched transformations
            unmatched_transforms: list[dict[str, any]] = [
                st for st in study_config.get('transformations', [])
                    if st.get('name') not in [rt.get('name') for rt in remote_transforms]
            ]
            unmatched_transform: dict[str, any]
            for unmatched_transform in unmatched_transforms:
                _logger.warning(
                    'Local transformations config entry "%s" (%s) not found in remote transformations config',
                    unmatched_transform.get('name'),
                    study_config.get('study')
                )

        # cache source => output field mappings for the enum/string properties for each transformation
        for study_config in self._study_configurations:
            transformation: dict[str, any]
            for transformation in study_config.get('transformations', []):
                sub_src_rec_enum_prop: str
                for sub_src_rec_enum_prop in self._sub_source_record_enum_properties:
                    self._sub_source_record_enum_properties[sub_src_rec_enum_prop] = (
                        self._find_source_field(transformation, sub_src_rec_enum_prop) or None
                    )

        return self._study_configurations

    def load_json_schema(self, save_local_copy: bool = False) -> dict[str, any]:
        """ Download JSON schema from configured URL """
        # download remote JSON schema file
        self._json_schema = json.loads(self._c3dc_file_manager.read_file(self._json_schema_url).decode('utf-8'))
        if save_local_copy:
            self._c3dc_file_manager.write_file(
                json.dumps(self._json_schema).encode('utf-8'),
                f'./{self._c3dc_file_manager.get_basename(self._json_schema_url)}'
            )

        if not self._json_schema:
            raise RuntimeError('Unable to get node types, JSON schema not loaded')
        if '$defs' not in self._json_schema:
            raise RuntimeError('Unable to get node types, JSON schema does not contain root-level "$defs" property')

        # log warnings for remote nodes not found in C3DC model
        unmapped_def: str
        for unmapped_def in [k for k in self._json_schema['$defs'] if k != 'nodes' and not C3dcEtlModelNode.get(k)]:
            _logger.warning('Schema "$defs" child node "%s" not defined in C3dcEtlModelNode enum', unmapped_def)

        # cache schema objects matching C3DC model nodes
        self._json_schema_nodes = {k:v for k,v in self._json_schema['$defs'].items() if C3dcEtlModelNode.get(k)}

        # cache allowed values for enum properties; map node.property => [permissible values]
        self._sub_source_record_enum_properties.clear()
        self._json_schema_property_enum_values.clear()
        node_type: str
        for node_type in self._json_schema_nodes:
            enum_properties: dict[str, any] = {
                k:v for k,v in self._get_json_schema_node_properties(node_type).items()
                    if 'enum' in v or 'enum' in v.get('items', {})
            }
            prop_name: str
            prop_props: dict[str, any]
            for prop_name, prop_props in enum_properties.items():
                enum_values: list[str] = prop_props.get('enum', None)
                enum_values = prop_props.get('items', {}).get('enum', []) if enum_values is None else enum_values
                self._json_schema_property_enum_values[f'{node_type}.{prop_name}'] = enum_values
                # enum properties that contain ';' in any of the permissible values don't support
                # multiple delimited values since ';' is the delimiting character
                if not any(C3dcEtl.MULTIPLE_VALUE_DELIMITER in v for v in enum_values):
                    self._sub_source_record_enum_properties[f'{node_type}.{prop_name}'] = None

        # cache allowed values for enum codes; map node.property => {enum code => enum value}
        self._json_schema_property_enum_code_values.clear()
        node_type_dot_property_name: str
        for node_type_dot_property_name, enum_values in self._json_schema_property_enum_values.items():
            self._json_schema_property_enum_code_values[node_type_dot_property_name] = {
                (v.partition(' : '))[0]:v for v in enum_values
            }

        return self._json_schema

    def validate_json_etl_data(self) -> None:
        """ Validate JSON ETL data resulting from study configurations specified in config """
        are_etl_data_sets_valid: bool = True
        study_configuration: dict[str, any]
        for study_configuration in self._study_configurations:
            transformation_name: str
            for transformation_name in [t.get('name') for t in study_configuration.get('transformations', [])]:
                are_etl_data_sets_valid = (
                    self._is_json_etl_data_valid(study_configuration.get('study'), transformation_name)
                    and
                    are_etl_data_sets_valid
                )
        return are_etl_data_sets_valid

    def create_json_etl_files(self) -> None:
        """ Create ETL files and save to output paths for study configurations specified in conf """
        study_configuration: dict[str, any]
        for study_configuration in self._study_configurations:
            transformation: dict[str, any]
            for transformation in study_configuration.get('transformations', []):
                self._create_json_etl_file(study_configuration.get('study'), transformation)

    def _verify_config(self) -> None:
        required_config_var: str
        missing_config_vars: list[str] = []
        for required_config_var in ('JSON_SCHEMA_URL', 'STUDY_CONFIGURATIONS'):
            if not self._config.get(required_config_var):
                missing_config_vars.append(required_config_var)
        if missing_config_vars:
            raise RuntimeError(
                f'One or more required variables not specified in configuration: {tuple(missing_config_vars)}'
            )

        study_configs: list[dict[str, any]] = json.loads(self._config.get('STUDY_CONFIGURATIONS', '[]'))
        study_config: dict[str, any]
        for study_config in study_configs:
            for required_config_var in ('study', 'transformations', 'transformations_url'):
                if not study_config.get(required_config_var):
                    missing_config_vars.append(f'STUDY_CONFIGURATIONS => {required_config_var}')
            xform: dict[str, any]
            for xform in study_config.get('transformations', []):
                for required_config_var in ('name', 'source_file_path', 'output_file_path'):
                    if not xform.get(required_config_var):
                        missing_config_vars.append(f'STUDY_CONFIGURATIONS => transformations => {required_config_var}')

        if missing_config_vars:
            raise RuntimeError(
                f'One or more required variables not specified in configuration: {tuple(missing_config_vars)}'
            )

    def _generate_uuid(self) -> uuid.UUID:
        """ Generate and return UUID(v4) using internal RNG that may be seeded for idempotent values """
        return uuid.UUID(int=self._random.getrandbits(128), version=4)

    def _get_race(self, source_race: str, source_ethnicity: str = None) -> list[str]:
        """ Determine race given specified source values of race and ethnicity """
        # Note delimited multi-entry source values are supported for both
        # race ("White;Other") and ethnicity ("Not Reported;Unknown")

        races: set[str] = set()
        # keep source ethnicity(-ies) if allowed and determinate race value like 'Hispanic or Latino' specified
        # and not indeterminate value such as 'Not Reported', 'Unknown', etc
        if source_ethnicity:
            source_ethnicities: set[str] = {
                e.strip() for e in source_ethnicity.split(C3dcEtl.MULTIPLE_VALUE_DELIMITER)
                    if e.strip().casefold() in {v.casefold() for v in C3dcEtl.ETHNICITY_ALLOWED_VALUES}
            }
            if all(
                e.casefold() in {v.casefold() for v in C3dcEtl.ETHNICITY_ALLOWED_VALUES} for e in source_ethnicities
            ):
                races.update(source_ethnicities)

        # keep source race(s) if:
        #   - specified as allowed and determinate value (not 'Unknown', 'Not Allowed to Collect', 'Not Reported', etc)
        #   *OR*
        #   - source ethnicity not specified or not allowed/determinate race value
        if source_race:
            source_races: set[str] = {r.strip() for r in source_race.split(C3dcEtl.MULTIPLE_VALUE_DELIMITER)}
            if not races:
                # no ethnicity value provided so just use race values whether they're determinate or not
                races.update(source_races)
            else:
                # determinate race value ('Hispanic or Latino') came through ethnicity field so add only determinate
                # values from race field since indeterminate values ('Not Reported', 'Unknown', etc) don't apply
                races.update(
                    {
                        r for r in source_races if r.strip().casefold() not in {
                            v.casefold() for v in C3dcEtl.RACE_UNDETERMINED_VALUES
                        }
                    }
                )

        return list(races)

    def _is_json_etl_data_valid(self, study_id: str, transformation_name: str) -> bool:
        """ Validate JSON ETL data for specified transformation against JSON schema """
        _logger.info('Validating JSON ETL data for transformation %s (study %s)', transformation_name, study_id)
        log_msg: str
        if not self._json_etl_data_sets.get(study_id, {}).get(transformation_name):
            _logger.warning('No JSON ETL data loaded to validate')
            return False

        if not self._json_schema:
            self.load_json_schema()
            if not self._json_schema:
                log_msg = f'Unable to download JSON schema from "{self._json_schema_url}" to perform validation'
                _logger.critical(log_msg)
                raise RuntimeError(log_msg)

        try:
            jsonschema.validate(
                instance=self._json_etl_data_sets[study_id][transformation_name],
                schema=self._json_schema
            )
            _logger.info(
                'Validation succeeded for JSON ETL data for transformation %s (study %s)',
                transformation_name,
                study_id
            )
            return True
        except jsonschema.exceptions.ValidationError as verr:
            _logger.error(
                'ETL data for transformation %s (study %s) failed schema validation:',
                transformation_name,
                study_id
            )
            _logger.error(verr)
            validator: jsonschema.Validator = jsonschema.Draft202012Validator(self._json_schema)
            validation_error: ValidationError
            for validation_error in validator.iter_errors(self._json_etl_data_sets[study_id][transformation_name]):
                _logger.error('%s: %s', validation_error.json_path, validation_error.message)
        return False

    def _save_json_etl_data(self, study_id: str, transformation: dict[str, any]) -> None:
        """ Save JSON ETL data for specified transformation to designated output file """
        _logger.info('Saving JSON ETL data to %s', transformation.get('output_file_path'))
        self._c3dc_file_manager.write_file(
            json.dumps(self._json_etl_data_sets[study_id][transformation.get('name')], indent=2).encode('utf-8'),
            transformation.get('output_file_path')
        )

    def _get_source_manifest_data(self, transformation: dict[str, any]) -> dict[dict[str, any]]:
        """ Read and return source file manifest data (guid, md5, etc) if available """
        manifests: dict[str, dict[str, any]] = {}
        if not (
            transformation.get('source_file_manifest_path')
            and
            self._c3dc_file_manager.file_exists(transformation.get('source_file_manifest_path'))
        ):
            _logger.info('Source file manifest data not available, skipping load')
            return {}

        _logger.info('Loading source file manifest data from "%s"', transformation.get('source_file_manifest_path'))
        manifest_file_path: str = transformation['source_file_manifest_path']
        if pathlib.Path(manifest_file_path).suffix.lower() != '.xlsx':
            raise RuntimeError(f'Unsupported source file manifest data file type: "{manifest_file_path}"')

        manifest_file_data: bytes = self._c3dc_file_manager.read_file(manifest_file_path)
        sheet_name: str = transformation.get('source_file_manifest_sheet', 'clinical_measure_file')[:31]
        wb: Workbook = openpyxl.load_workbook(io.BytesIO(manifest_file_data), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f'Worksheet "{sheet_name}" not found in workbook worksheet list: {wb.sheetnames}')
        ws: Worksheet = wb[sheet_name] if sheet_name else wb.worksheets[0]
        row: str
        cols: list[str] = []
        manifest_fields: dict[str, str] = {
            'participant.participant_id': 'participant_id',
            'file_name': 'file_name',
            'dcf_indexd_guid': 'guid',
            'md5sum': 'md5',
            'file_size': 'size',
            'file_url': 'url'
        }
        for row in ws.iter_rows():
            if not cols:
                # first row is header
                cols = [cell.value for cell in row]
                continue
            manifest: dict[str, any] = {}
            col_num: int
            for col_num, cell in [(i,v) for i,v in enumerate(row) if cols[i] in manifest_fields]:
                value: any = cell.value
                manifest[manifest_fields[cols[col_num]]] = str(value).strip() if value is not None else ''
            if (
                manifest['participant_id']
                and
                manifest['file_name']
                and
                f'{manifest["participant_id"]}.json' == manifest['file_name']
                and any(v not in ('', None, -1) for v in manifest.values())
            ):
                if manifest['file_name'] in manifests:
                    raise RuntimeError(f'Duplicate file name found in manifest sheet: "{manifest["file_name"]}"')
                manifests[manifest['file_name']] = manifest
        if not manifests:
            raise RuntimeError('No manifest records loaded')
        return manifests

    def _get_row_mapped_node_mappings(
        self,
        transformation: dict[str, any],
        node: C3dcEtlModelNode
    ) -> list[dict[str, any]]:
        """ Read and return row-mapped node mappings from file and sheet specified in transformation """
        mappings: list[dict[str, any]] = []
        mappings_path_config_name: str = f'{node}_mappings_path'
        mappings_sheet_config_name: str = f'{node}_mappings_sheet'
        if not (
            transformation.get(mappings_path_config_name)
            and
            self._c3dc_file_manager.file_exists(transformation.get(mappings_path_config_name))
        ):
            _logger.info('Mappings file for "%s" not available, skipping load', node)
            return []

        _logger.info('Loading "%s" mappings from "%s"', node, transformation.get(mappings_path_config_name))
        mappings_file_path: str = transformation[mappings_path_config_name]
        if pathlib.Path(mappings_file_path).suffix.lower() != '.xlsx':
            raise RuntimeError(f'Unsupported "{node}" mappings file type: "{mappings_file_path}"')

        mappings_file_data: bytes = self._c3dc_file_manager.read_file(mappings_file_path)
        sheet_name: str = transformation.get(mappings_sheet_config_name, f'phs002790_{node}')[:31]
        wb: Workbook = openpyxl.load_workbook(io.BytesIO(mappings_file_data), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f'Worksheet "{sheet_name}" not found in workbook worksheet list: {wb.sheetnames}')
        ws: Worksheet = wb[sheet_name] if sheet_name else wb.worksheets[0]
        rownum: int
        row: any
        cols: list[str] = []
        for rownum, row in enumerate(ws.iter_rows(), start=1):
            if not cols:
                # first row is header
                cols = [cell.value for cell in row]
                if len(cols) != len(set(cols)):
                    raise RuntimeError('Total columns > unique columns; check for duplicate columns')
                continue
            mapping: dict[str, any] = {}
            col_num: int
            for col_num, cell in enumerate(row):
                value: any = cell.value
                mapping[cols[col_num]] = str(value).strip() if value is not None else ''
            if not mapping or all(m in ('', None) for m in mapping):
                _logger.warning('Invalid/empty mapping in row %d, skipping', rownum)
                continue
            if any(mapping == m for m in mappings):
                _logger.warning('Duplicate mapping in row %d, skipping', rownum)
                continue
            mappings.append(mapping)
        if not mappings:
            raise RuntimeError(f'No "{node}" mapping records loaded')
        return mappings

    def _get_primary_follow_up(self, upi: str, follow_ups: list[list[dict[str, any]]]) -> list[dict[str, any]]:
        """ Return first/earliest 'Dead' or last/latest 'Alive' follow-up in list of data lists """
        if not follow_ups:
            raise RuntimeError(f'Unable to determine primary follow-up for subject "{upi}", no follow-ups provided')

        # return first/earliest 'Dead' follow up if any
        dead_follow_ups: list[list[dict[str, any]]] = [
            dl for dl in follow_ups if any(
                (d.get('form_field_id') or '').upper().strip() == 'PT_VST'
                and
                (d.get('value') or '').upper().strip() == 'DEAD'
                for d in dl
            )
        ]
        sorted_dead_follow_ups: list[list[dict[str, any]]] = sorted(
            dead_follow_ups,
            key=lambda dl: max(
                (int(d.get('value', 0)) for d in dl if (d.get('form_field_id') or '').upper().strip() == 'PT_FU_BEGDT'),
                default=0
            ),
            reverse=False
        )
        if sorted_dead_follow_ups:
            return sorted_dead_follow_ups[0]

        # return last/latest 'Alive' follow up if any
        alive_follow_ups: list[list[dict[str, any]]] = [
            dl for dl in follow_ups if any(
                (d.get('form_field_id') or '').upper().strip() == 'PT_VST'
                and
                (d.get('value') or '').upper().strip() == 'ALIVE'
                for d in dl
            )
        ]
        sorted_alive_follow_ups: list[list[dict[str, any]]] = sorted(
            alive_follow_ups,
            key=lambda dl: max(
                (int(d.get('value', 0)) for d in dl if (d.get('form_field_id') or '').upper().strip() == 'PT_FU_BEGDT'),
                default=0
            ),
            reverse=True
        )
        if sorted_alive_follow_ups:
            return sorted_alive_follow_ups[0]

        # no follow ups with vital status 'Dead' or 'Alive', source data inspection and/or code adjustment is needed
        raise RuntimeError(
            f'Unable to determine primary follow-up for subject "{upi}", no follow-ups ' +
            'provided having vital status ("PT_VST") "Alive" or "Dead"')

    def _get_primary_source_form_data_list(self, upi: str, form: dict[str, any]) -> list[dict[str, any]]:
        """ Find and return highest priority 'data' element of specified source record form object """
        data_lists: list[list[dict[str, any]]] = form.get('data', [])
        if not data_lists:
            raise RuntimeError(f'Form "data" element for subject "{upi}" is missing or empty: {form}')
        if len(data_lists) == 1:
            return data_lists[0]

        match form.get('form_id'):
            case 'FOLLOW_UP':
                return self._get_primary_follow_up(upi, data_lists)
            case _:
                # form has multiple (duplicate in original source JSON) 'data' elements and determination of primary
                # element hasn't been implemented (only 'FOLLOW_UP' handled for now), but don't abort processing unless
                # form has one or more field(s) that need to be retrieved and therefore are mapped
                form_field_ids: set[str] = {d.get('form_field_id') for dl in data_lists for d in dl}
                if form_field_ids.intersection(self._source_field_output_types.keys()):
                    raise RuntimeError(
                        f'Unable to determine primary data list for form "{form.get("form_id")}" containing multiple ' +
                        f'"data" elements for subject "{upi}"'
                    )
                _logger.warning(
                    (
                        'Unable to determine primary data list for form "%s" containing multiple "data" elements ' +
                        'for subject "%s", returning first data list in source file'
                    ),
                    form.get('form_id'),
                    upi
                )
                return data_lists[0]

    def _load_source_data(
        self,
        study_id: str,
        transformation: dict[str, any],
        force_reload: bool = False
    ) -> list[dict[str, any]]:
        """ Load raw ETL data from source path specified in config """
        msg: str
        recs: list[dict[str, any]] = []
        rec: dict[str, any]

        if not force_reload and self._raw_etl_data_objects.get(study_id, {}).get(transformation.get('name')):
            _logger.info('Source data already loaded, skipping load')
            return self._raw_etl_data_objects[study_id][transformation.get('name')]

        _logger.info('Loading source data')

        # load source file metadata from manifest file
        manifests: dict[str, dict[str, any]] = self._get_source_manifest_data(transformation)

        # load source data from individual json files
        source_file_parent_location: str = transformation.get('source_file_path')
        source_file_locations: list[str] = sorted(
            f for f in self._c3dc_file_manager.list_files(source_file_parent_location) if f.endswith('.json')
        )
        source_file_location: str
        processed: int = 0
        for source_file_location in source_file_locations:
            source_file_name: str = C3dcFileManager.get_basename(source_file_location)
            processed += 1
            if processed % 100 == 0:
                _logger.info('%d of %d source file(s) processed', processed, len(source_file_locations))
            # construct minimal object from raw source data limited to source fields specified in mappings
            obj: any = json.loads(
                self._c3dc_file_manager.read_file(source_file_location).decode('utf-8'),
                object_pairs_hook=C3dcEtl.collate_form_data
            )
            rec = {'source_file_name': source_file_name, 'manifest': manifests.get(source_file_name, {})}

            source_field_name: str
            source_field_output_type: str
            # retrieve root-level properties such as 'upi'
            for source_field_name, source_field_output_type in {
                k:v for k,v in self._source_field_output_types.items() if k in obj
            }.items():
                if source_field_name in rec and source_field_output_type != 'array':
                    msg = f'Duplicate source field "{source_field_name}" found in file "{source_file_name}"'
                    _logger.fatal(msg)
                    raise RuntimeError(msg)
                rec[source_field_name] = obj.get(source_field_name)

            if 'upi' not in rec:
                raise RuntimeError(f'Subject identifier ("upi") not specified for source file "{source_file_location}"')

            # get properties such as DEMOGRAPHY=>DM_BRTHDAT that are defined within forms
            form_id: str
            data_fields: list[dict[str, any]]
            for form_id, data_fields in {
                f.get('form_id'):self._get_primary_source_form_data_list(rec['upi'], f) for f in obj.get('forms', [])
            }.items():
                data_field: dict[str, any]
                for data_field in [
                    d for d in data_fields
                        if d.get('form_field_id') in self._source_field_output_types or
                            f'{form_id}.{d.get("form_field_id")}' in self._source_field_output_types
                ]:
                    source_field_name = data_field.get('form_field_id')
                    form_id_field_id = f'{form_id}.{source_field_name}'
                    # use the full form-qualified name if specified in mapping (e.g. to avoid dupes)
                    if form_id_field_id in self._source_field_output_types:
                        source_field_name = form_id_field_id
                    source_field_output_type = self._source_field_output_types[source_field_name]

                    if source_field_output_type == 'array':
                        rec[source_field_name] = rec.get(source_field_name, [])
                        rec[source_field_name].append(data_field.get('value'))
                    elif source_field_name in rec:
                        msg = (
                            f'Duplicate source field "{source_field_name}" ({form_id_field_id}) ' +
                            f'found in file "{source_file_name}"'
                        )
                        _logger.fatal(msg)
                        raise RuntimeError(msg)
                    else:
                        rec[source_field_name] = data_field.get('value')
            recs.append(rec)

        self._raw_etl_data_objects[study_id] = self._raw_etl_data_objects.get(study_id, {})
        self._raw_etl_data_objects[study_id][transformation.get('name')] = recs
        return self._raw_etl_data_objects[study_id][transformation.get('name')]

    def _load_study_data(self, study_configuration: dict[str, any]) -> dict[str, list[dict[str, any]]]:
        """ Load raw ETL data from source files specified in study config """
        transformation: dict[str, any]
        for transformation in study_configuration.get('transformations', []):
            self._load_source_data(study_configuration.get('study'), transformation)
        return self._raw_etl_data_objects[study_configuration.get('study')]

    def _get_json_schema_node_property_type(self, node_type_dot_property_name: str) -> str:
        """ Get JSON schema property type for specified output field ('node_name.field_name') """
        if '.' not in node_type_dot_property_name:
            raise RuntimeError(f'Invalid node.property specified: {node_type_dot_property_name}')
        node_type: str = node_type_dot_property_name.split('.')[0]
        property_name: str = node_type_dot_property_name.split('.')[-1]
        return self._json_schema_nodes.get(node_type, {}).get('properties', {}).get(property_name, {}).get('type')

    def _get_json_schema_node_properties(self, node_type: C3dcEtlModelNode) -> dict[str, any]:
        """ Get properties for specified node in JSON schema  """
        if not self._json_schema_nodes:
            raise RuntimeError('Unable to get JSON schema nodes')
        if node_type not in self._json_schema_nodes:
            raise RuntimeError(f'Unable to get JSON schema node for type "{node_type}"')
        if 'properties' not in self._json_schema_nodes[node_type]:
            raise RuntimeError(f'"properties" not found in JSON schema node for type "{node_type}"')
        return self._json_schema_nodes[node_type]['properties']

    def _get_json_schema_node_required_properties(self, node_type: C3dcEtlModelNode) -> dict[str, any]:
        """ Get required properties for specified node in JSON schema  """
        node_properties: dict[str, any] = self._get_json_schema_node_properties(node_type)
        return {
            k:v for k,v in node_properties.items() if self._is_json_schema_node_property_required(f'{node_type}.{k}')
        }

    def _get_json_schema_node_property_converted_value(
        self,
        node_type_dot_property_name: str,
        value: any
    ) -> list | int | str | None:
        """ Get output value converted to JSON schema type for specified property array, integer, string """
        if value is None:
            return None

        # collate into string if collection specified
        value = C3dcEtl.MULTIPLE_VALUE_DELIMITER.join(value) if isinstance(value, (list, set, tuple)) else value

        if '.' not in node_type_dot_property_name:
            raise RuntimeError(f'Unexpected schema property name ("." not present): "{node_type_dot_property_name}"')

        json_type: str = self._get_json_schema_node_property_type(node_type_dot_property_name)
        if json_type not in C3dcEtl.TYPE_NAME_CLASS_MAP:
            raise RuntimeError(f'Schema type "{json_type}" not in type name class map')

        python_type = C3dcEtl.TYPE_NAME_CLASS_MAP.get(json_type)
        if python_type == list:
            if node_type_dot_property_name not in self._json_schema_property_enum_values:
                # values not constrained, split on delimiter
                return [s.strip() for s in str(value).split(C3dcEtl.MULTIPLE_VALUE_DELIMITER)]

            # multi-valued properties may specify multiple values using delimiter
            vals: dict[str, str] = {v:v for v in value.split(C3dcEtl.MULTIPLE_VALUE_DELIMITER)}
            val: str
            for val in vals:
                case_matched_val: str = self._case_match_json_schema_enum_value(node_type_dot_property_name, val)
                if case_matched_val is not None:
                    vals[val] = case_matched_val
                else:
                    _logger.warning(
                        'Unable to case match source sub-value "%s" for enum property "%s", omitting',
                        val,
                        node_type_dot_property_name
                    )
            return [v for v in vals.values() if v is not None]
        if python_type == int:
            return int(float(value))
        if python_type == str:
            if node_type_dot_property_name not in self._json_schema_property_enum_values:
                return str(value)
            new_value: str = self._case_match_json_schema_enum_value(node_type_dot_property_name, str(value))
            if new_value is None:
                _logger.warning(
                    'Unable to case match source value "%s" for enum property "%s", omitting',
                    value,
                    node_type_dot_property_name
                )
            return new_value

        raise RuntimeError(f'Unsupported output value conversion type: "{python_type}" (schema type "{json_type}")')

    def _case_match_json_schema_enum_value(self, node_type_dot_property_name: str, value: any) -> any:
        """ Align case of specified enum value with JSON schema permissible values, e.g. 'unknown' => 'Unknown' """
        if value is None or not isinstance(value, str):
            return value

        enum_values: list[str] = self._json_schema_property_enum_values.get(node_type_dot_property_name)
        if not enum_values:
            return value

        enum_value_matches: list[str] = [e for e in enum_values if e.casefold() == str(value).casefold()]
        if len(enum_value_matches) > 1:
            raise RuntimeError(
                f'Multiple enum value matches for "{value}" found in schema property "{node_type_dot_property_name}"'
            )
        return enum_value_matches[0] if len(enum_value_matches) == 1 else None

    def _is_json_schema_node_property_required(self, node_type_dot_property_name: str) -> bool:
        """ Determine if JSON schema property for specified output field ('node_name.field_name') is required """
        if '.' not in node_type_dot_property_name:
            raise RuntimeError(f'Invalid node.property specified: {node_type_dot_property_name}')
        node_type: str = node_type_dot_property_name.split('.')[0]
        property_name: str = node_type_dot_property_name.split('.')[-1]
        return property_name in self._json_schema_nodes.get(node_type, {}).get('required', [])

    def _get_replacement_new_value_errors(
        self,
        source_header: list[str],
        study_id: str,
        transformation_name: str,
        replacement_new_value: any
    ) -> list[str]:
        """
        Get errors for specified transformation mapping replacement value (new_value), ignoring allowed
        values such as for phs_accession (literal strings within square brackets e.g. "['phs123456']").
        Errors may occur in specification of macro-like substitutions contained within square brackets as below:
        [uuid] => substitute with a UUID (v4)
        [field:{source field name}] => substitute with specified record's source field value, e.g. [field:TARGET USI]
        """
        if '[' not in str(replacement_new_value) and ']' not in str(replacement_new_value):
            return []

        errors: list[str] = []
        macro: str
        for macro in re.findall(r'\{.*?\}', replacement_new_value):
            macro_text: str = macro.strip(' {}').strip()
            # source field to be replaced will be specified as '{field: FIELD_NAME}'
            if not (
                (macro_text.startswith('"') and macro_text.endswith('"')) or
                (macro_text.startswith("'") and macro_text.endswith("'")) or
                macro_text.lower() in ('find_enum_value', 'race', 'sum', 'sum_abs_first', 'uuid') or
                (
                    macro_text.lower().startswith('field:')
                    and
                    macro_text[len('field:'):] in source_header
                )
            ):
                errors.append(f'{transformation_name} ({study_id}): invalid replacement macro: {replacement_new_value}')

        return errors

    def _get_transformation_mapping_errors(
        self,
        study_id: str,
        transformation_name: str,
        mapping: dict[str, any]
    ) -> list[str]:
        """ Get errors for specified transformation mapping """
        errors: list[str] = []

        compound_source_fields: list[str] = []
        source_header: list[str] = list(self._source_field_output_types.keys())
        source_field: str = mapping.get('source_field')
        if not source_field:
            errors.append(f'{transformation_name} ({study_id}): mapping source field not specified: {mapping}')
        if source_field.startswith('[') and source_field.endswith(']'):
            # strip extra spaces; csv module parses "field 1, field 2" into ["field 1", " field 2"]
            compound_source_fields = [s.strip() for s in next(csv.reader([source_field.strip(' []')]))]
            if not {s for s in compound_source_fields if s != 'string_literal'}.issubset(set(source_header)):
                errors.append(
                    f'{transformation_name} ({study_id}): compound source field in mapping ("{source_field}") ' +
                    f'not present in source data header: {source_header}'
                )
                errors.append('Verify all mapped source fields in remote transformations file are in source data')
        elif source_field not in source_header:
            errors.append(
                (
                    f'{transformation_name} ({study_id}): source field in mapping ("{source_field}") ' +
                    f'not present in source data header ("{source_header}")'
                )
            )
            errors.append('Verify all mapped source fields in remote transformations file are in source data')
        output_field: str = mapping.get('output_field')
        if not output_field:
            errors.append(f'{transformation_name} ({study_id}): mapping output field not specified: {mapping}')
        output_field_parts = output_field.split('.')
        output_node: str = output_field_parts.pop(0)
        output_property: str = '.'.join(output_field_parts)
        if (
            output_node not in self._json_schema_nodes
            or
            output_property not in self._json_schema_nodes.get(output_node, {}).get('properties', {})
        ):
            errors.append(f'{transformation_name} ({study_id}): mapping output field invalid: {mapping}')

        replacement_entry: dict[str, str]
        for replacement_entry in mapping.get('replacement_values', []):
            if 'old_value' not in replacement_entry or 'new_value' not in replacement_entry:
                errors.append(
                    f'{transformation_name} ({study_id}): replacement entry missing new or old value: ' +
                    str(replacement_entry)
                )

            old_value: str = replacement_entry.get('old_value', '*')
            new_value: any = replacement_entry.get('new_value', '')

            if source_field == '[string_literal]' and old_value not in ('+', '*'):
                errors.append(
                    f'{transformation_name} ({study_id}): replacement entry has invalid old value for ' +
                    f'string literal source: {mapping}'
                )

            # if multiple source fields then old value must be wildcard or be multi-valued with matching count
            if old_value != '*' and compound_source_fields:
                old_values: list[str] = [
                    o.strip() for o in next(
                        csv.reader([old_value.strip(' []')], delimiter=C3dcEtl.MULTIPLE_VALUE_DELIMITER)
                    )
                ]
                if len(old_values) != len(compound_source_fields):
                    errors.append(
                        f'{transformation_name} ({study_id}): replacement entry for mapping with multiple source ' +
                        'field (macro?) must have old value set to wildcard ("*") or contain same number of values ' +
                        'as source fields, delimited by "{C3dcEtl.MULTIPLE_VALUE_DELIMITER}"'
                    )
                    errors.append('ex: [source_field_1,source_field_2] => "old_value_1;old_value_2"')

            # new_value can be list or scalar so handle scalars as single-valued lists and enumerate all
            new_value = [new_value] if not isinstance(new_value, (list, set, tuple)) else new_value
            new_sub_value: str
            for new_sub_value in new_value:
                errors.extend(
                    self._get_replacement_new_value_errors(source_header, study_id, transformation_name, new_sub_value)
                )

        return errors

    def _get_transformation_errors(self, study_id: str, index: int, transformation: dict[str, any]) -> list[str]:
        """ Get errors for specified transformation """
        errors: list[str] = []
        required_properties: tuple[str, ...] = ('name', 'source_file_path', 'output_file_path', 'mappings')
        if any(not transformation.get(p) for p in required_properties):
            errors.append(
                f'Study {study_id}, transformation {index + 1}: one or more of "{required_properties}" missing/invalid'
            )

        if (
            transformation.get('source_file_path')
            and
            not self._c3dc_file_manager.file_exists(transformation.get('source_file_path'))
        ):
            errors.append(
                f'{transformation.get("name")} ({study_id}): invalid source file parent dir ' +
                f'"{transformation.get("source_file_path")}"'
            )

        if (
            transformation.get('name')
            and
            not (self._raw_etl_data_objects or {}).get(study_id, {}).get(transformation.get('name'))
        ):
            if (
                transformation.get('source_file_path')
                and
                self._c3dc_file_manager.file_exists(transformation.get('source_file_path'))
            ):
                self._load_source_data(study_id, transformation)
            if not (self._raw_etl_data_objects or {}).get(study_id, {}).get(transformation.get('name')):
                errors.append(f'{transformation.get("name")}: unable to load source data')

        if not errors:
            mapping: dict[str, any]
            for mapping in transformation['mappings']:
                mapping_errors: list[str] = self._get_transformation_mapping_errors(
                    study_id, transformation.get('name'), mapping
                )
                errors.extend(mapping_errors)
        return errors

    def _get_study_configuration_errors(self, index: int, study_configuration: dict[str, any]) -> list[str]:
        """ Get errors for specified study configuration """
        errors: list[str] = []
        required_properties: tuple[str, ...] = ('study', 'transformations_url', 'transformations')
        if any(not study_configuration.get(p) for p in required_properties):
            errors.append(f'Study configuration {index + 1}: one or more of "{required_properties}" missing/invalid')
        t_index: int
        transformation: dict[str, any]
        for t_index, transformation in enumerate(study_configuration.get('transformations')):
            errors.extend(self._get_transformation_errors(study_configuration.get('study'), t_index, transformation))
        return errors

    def _assert_valid_study_configurations(self) -> None:
        """ Assert that study configurations specified in config are invalid else raise exception(s) """
        _logger.info('Validating study configurations')
        if not self._study_configurations:
            raise RuntimeError('No study configurations to validate')

        errors: list[str] = []

        if len(self._study_configurations) != len({st.get('study', '') for st in self._study_configurations}):
            raise RuntimeError('Duplicate study ids found in study configurations')

        sc_index: int
        study_configuration: dict[str, any]
        for sc_index, study_configuration in enumerate(self._study_configurations):
            errors.extend(self._get_study_configuration_errors(sc_index, study_configuration))

        error: str
        for error in errors:
            if 'PYTEST_CURRENT_TEST' in os.environ:
                print(error)
            _logger.error(error)
        if errors:
            raise RuntimeError('Invalid transformation(s) found')

    def _build_reference_file_fieldset_mapping(
        self,
        transformation: dict[str, any],
        participant: dict[str, any],
        source_file_manifest: dict[str, any],
        type_group_index: int
    ) -> list[dict[str, any]]:
        """
        Build reference file JSON transformation mapping for specified
        participant, transformation, and type group index
        """
        usi: str = participant.get('participant_id')
        if not usi:
            msg = 'Participant record missing "participant_id", unable to build reference file mapping'
            _logger.critical(msg)
            _logger.critical(participant)
            raise RuntimeError(msg)

        dcf_indexd_guid: str = source_file_manifest.get('guid')

        source_file_name: str = f'{usi}.json'
        source_file_parent_path: str = transformation.get('source_file_path')
        source_file_path: str = C3dcFileManager.join_location_paths(source_file_parent_path, source_file_name)
        if not self._c3dc_file_manager.file_exists(source_file_path):
            msg = 'Source file "{source_file_name}" not found'
            _logger.critical(msg)
            raise RuntimeError(msg)

        manifest_key: str
        for manifest_key in ('md5', 'size', 'url'):
            if source_file_manifest.get(manifest_key) in (None, ''):
                _logger.warning('No "%s" value specified in manifest for participant "%s"', manifest_key, usi)
        md5sum: str = source_file_manifest.get('md5')
        if not md5sum:
            # md5sum not specified in source file manifest, calculate manually
            md5sum = hashlib.md5(self._c3dc_file_manager.read_file(source_file_path)).hexdigest()
        file_size: int = int(source_file_manifest.get('size', -1))
        if file_size < 0:
            file_size = self._c3dc_file_manager.get_file_size(source_file_path)
        url: str = source_file_manifest.get('url', '')

        output_field_new_values: dict[str, any] = {
            'reference_file.dcf_indexd_guid': dcf_indexd_guid,
            'reference_file.file_name': source_file_name,
            'reference_file.file_type': 'json',
            'reference_file.file_category': 'input source data',
            'reference_file.file_size': file_size,
            'reference_file.md5sum': md5sum,
            'reference_file.file_description': 'JSON file containing input source data',
            'reference_file.reference_file_url': url
        }

        reference_file_fieldset: list[dict[str, any]] = []

        source_field: str = '[string_literal]'
        type_group_index_str: str = str(type_group_index)
        old_value: str = '*'
        output_field: str
        new_value: str | int
        for output_field, new_value in output_field_new_values.items():
            replacement_values: list[dict[str, any]] = [{
                'old_value': old_value,
                'new_value': new_value
            }]

            reference_file_field: dict[str, any] = {
                'output_field': output_field,
                'source_field': source_field,
                'type_group_index': type_group_index_str,
                'replacement_values': replacement_values
            }
            reference_file_fieldset.append(reference_file_field)
        return reference_file_fieldset

    def _build_reference_file_fieldset_mappings(
        self,
        transformation: dict[str, any],
        participants: list[dict[str, any]],
        source_file_manifests: list[dict[str, any]]
    ) -> dict[str, any]:
        """
        Source data consists of individual JSON files per subject (SUBJECT1.json, SUBJECT2.json, etc) and
        not all will make it into the harmonized data output file due to properties required by output schema 
        being invalid/missing/null. This method will dynamically populate the necessary reference file mapping
        for the subject (participant) records that passed validation and were loaded by the ETL. The reference
        file mappings will then be processed, populating the actual reference file entries in the harmonized
        data output file
        """
        # determine starting type group index for input source file reference mappings; should be after
        # existing reference file entries (programmatic source code, schema, transformation/mapping)
        type_group_index: int = max(
            int(m.get('type_group_index', 0)) for m in transformation.get('mappings', [])
                if m.get('output_field', '') == 'reference_file.file_name'
        )
        if type_group_index == 1:
            raise RuntimeError('Unexpected reference file type group index value: 1')

        reference_file_fieldset_mappings: list[dict[str, any]] = []
        participant: dict[str, any]
        processed: int = 0
        for participant in participants:
            processed += 1
            if processed % 100 == 0:
                _logger.info(
                    '%d of %d participant reference file fieldset mappings built',
                    processed,
                    len(participants)
                )
            source_file_manifest: dict[str, any] = [
                m for m in source_file_manifests if m.get('file_name') == f'{participant.get("participant_id")}.json'
            ]
            source_file_manifest = source_file_manifest[0] if source_file_manifest else {}
            type_group_index += 1
            reference_file_fieldset_mapping: list[dict[str, any]] = self._build_reference_file_fieldset_mapping(
                transformation,
                participant,
                source_file_manifest,
                type_group_index
            )
            reference_file_fieldset_mappings.extend(reference_file_fieldset_mapping)
        return reference_file_fieldset_mappings

    def _append_reference_file_fieldset_mappings(
        self,
        study_id: str,
        transformation: dict[str, any],
        participants: list[dict[str, any]]
    ) -> None:
        """ Append reference file mappings to specified transformation for specified participants """
        if any(
            m.get('output_field') == f'{C3dcEtlModelNode.REFERENCE_FILE}.file_category' and
                any(r.get('new_value') == 'input source data' for r in m.get('replacement_values', []))
            for m in transformation.get('mappings', [])
        ):
            # transformation mapping already has 'input source data' reference file entries, skip
            _logger.warning(
                'Existing "input source data" reference file transformation mapping entries found, ' +
                'skipping construction of reference file fieldset mappings'
            )
            return

        _logger.info('Building reference file fieldset mappings for source files')
        source_file_manifests: list[dict[str, any]] = [
            m.get('manifest') for m in
                self._raw_etl_data_objects.get(study_id, {}).get(transformation.get('name'), {})
                    if m.get('manifest')
        ]
        reference_file_fieldset_mappings: list[dict[str, any]] = self._build_reference_file_fieldset_mappings(
            transformation,
            participants,
            source_file_manifests
        )
        transformation.get('mappings').extend(reference_file_fieldset_mappings)
        _logger.info(
            'Built and appended %d "%s" mapping entries for %d "%s" records',
            len(reference_file_fieldset_mappings),
            C3dcEtlModelNode.REFERENCE_FILE,
            len(participants),
            C3dcEtlModelNode.PARTICIPANT
        )
        study_config: dict[str, any] = [c for c in self._study_configurations if c.get('study') == study_id]
        study_config = study_config[0] if study_config else {}
        remote_config: dict[str, any] = json.loads(
            self._c3dc_file_manager.read_file(study_config.get('transformations_url')).decode('utf-8')
        )
        remote_transform: dict[str, any] = [
            t for t in remote_config.get('transformations') if t.get('name') == transformation.get('name')
        ]
        remote_transform = remote_transform[0] if remote_transform else {}
        remote_transform.get('mappings').extend(reference_file_fieldset_mappings)

        # save updated transformation
        save_path_components: list[str] = list(
            C3dcFileManager.split_location_paths(study_config.get("transformations_url"))
        )
        orig_save_path: pathlib.Path = pathlib.Path(
            C3dcFileManager.get_basename(study_config.get("transformations_url"))
        )
        save_path_components[-1] = f'{orig_save_path.stem}.ref_files{orig_save_path.suffix}'
        save_path: str = C3dcFileManager.join_location_paths(*save_path_components)
        self._c3dc_file_manager.write_file(json.dumps(remote_config, indent=4).encode('utf-8'), save_path)
        _logger.info('Saved updated remote transformation mapping to %s', save_path)

    def _get_mapped_output_value(
        self,
        mapping: dict[str, any],
        source_record: dict[str, any]
    ) -> any:
        """
        Get output value for specified mapping and source record.
        
        Mappings can specify old values to be replaced that are explicit values (replace on exact match) or
        wildcards using '+' or '*', where '+' results in replacement only if there is an existing value and
        '*' always results in replacement, whether the existing value is populated or null/blank. The new value
        can be specified with explicit scalar values or macro-like substitution directives as specified below:
        {uuid} => substitute with a UUID (v4 w/ optional seed value for RNG from config)
        {sum} => substitute with sum of values for specified fields in 'source_field' propery
        {field:source_field_name} => substitute with specified record's source field value, e.g. {field:TARGET USI}
        """
        msg: str
        output_field: str = mapping.get('output_field')
        output_value: any = None

        source_field: str = mapping.get('source_field').strip(' \'"')
        source_value: str = source_record.get(source_field, None)

        default_value: any = mapping.get('default_value', None)

        replacement_entry: dict[str, str]
        for replacement_entry in mapping.get('replacement_values', []):
            old_value: str = replacement_entry.get('old_value', '*')
            new_value: any = replacement_entry.get('new_value', None)

            # check for and apply macros such as 'sum' if specified for new value
            new_vals: list[any] = new_value if isinstance(new_value, (list, set, tuple)) else [new_value]
            new_val: any
            for i, new_val in enumerate(new_vals):
                if not (str(new_val).startswith('{') and str(new_val).endswith('}')):
                    continue
                macros: list[str] = re.findall(r'\{.*?\}', new_val)
                if not macros:
                    continue
                macro: str = macros[0]
                macro_text: str = macro.strip(' {}').strip()
                if macro_text.lower() == 'uuid':
                    new_val = new_val.replace(macro, str(self._generate_uuid()))
                elif macro_text.lower().startswith('field:'):
                    # source field to be replaced will be specified as '[field: FIELD_NAME]
                    macro_field: str = macro_text[len('field:'):].strip()
                    if macro_field not in source_record:
                        msg = f'Macro field not found in source record: "{macro_field}"'
                        _logger.critical(msg)
                        raise RuntimeError(msg)
                    new_val = new_val.replace(macro, source_record[macro_field])
                elif macro_text.lower() == 'find_enum_value':
                    # source field will be code such as '8000/0' or 'C71.9' that can be found in output value enum
                    enum_value: any = self._get_json_schema_node_property_converted_value(
                        output_field,
                        self._json_schema_property_enum_code_values.get(output_field, {}).get(source_value)
                    )
                    if source_value and not enum_value:
                        _logger.warning('No enum value found for "%s" value code "%s"', source_field, source_value)
                    new_val = enum_value
                elif macro_text.lower() == 'laterality':
                    # source field should contain list of source fields to be added together
                    if not (source_field.startswith('[') and source_field.endswith(']')):
                        msg = (
                            f'Invalid source field "{source_field}" for "{macro_text.lower()}" macro in source file ' +
                            f'{source_record["source_file_name"]}, must be comma-delimited ' +
                            '(csv) string within square brackets, e.g. "[field1, field2]"'
                        )
                        _logger.critical(msg)
                        raise RuntimeError(msg)
                    # strip extra spaces; csv module parses "field 1, field 2" into ["field 1", " field 2"]
                    source_field_names: list[str] = [s.strip() for s in next(csv.reader([source_field.strip(' []')]))]
                    laterality: str = default_value
                    source_field_name: str
                    for source_field_name in source_field_names:
                        src_val: str = source_record.get(source_field_name)
                        enum_value: any = self._get_json_schema_node_property_converted_value(
                            output_field,
                            src_val
                        )
                        if source_value and not enum_value:
                            _logger.warning(
                                'No enum value found for laterality source value "%s" (source field "%s")',
                                src_val,
                                source_field_name
                            )
                        if enum_value:
                            laterality = enum_value
                            break
                    new_val = laterality
                elif macro_text.lower() in ('sum', 'sum_abs_first'):
                    # source field should contain list of source fields to be added together
                    if not (source_field.startswith('[') and source_field.endswith(']')):
                        msg = (
                            f'Invalid source field "{source_field}" for "{macro_text.lower()}" macro in source file ' +
                            f'{source_record["source_file_name"]}, must be comma-delimited ' +
                            '(csv) string within square brackets, e.g. "[field1, field2]"'
                        )
                        _logger.critical(msg)
                        raise RuntimeError(msg)
                    # strip extra spaces; csv module parses "field 1, field 2" into ["field 1", " field 2"]
                    source_field_names: list[str] = [s.strip() for s in next(csv.reader([source_field.strip(' []')]))]
                    addends: list[float | int] = []
                    source_field_name: str
                    for source_field_name in source_field_names:
                        addend: str = source_record.get(source_field_name)
                        addend = '' if addend is None else str(addend).strip()
                        if addend in (None, ''):
                            # set output sum to blank/null if any addend is invalid/blank/null
                            return None

                        if not C3dcEtl.is_number(addend):
                            msg = (
                                f'Invalid "{source_field_name}" value "{addend}" for "{macro_text}" macro in source ' +
                                f'file {source_record["source_file_name"]}, must be a number'
                            )
                            _logger.warning(msg)
                            addends.append(None)
                        else:
                            addend = float(addend)
                            # use absolute value of first addend for 'sum_abs_first' macro
                            addend = abs(addend) if macro_text.lower() == 'sum_abs_first' and not addends else addend
                            addends.append(addend if not addend.is_integer() else int(addend))
                    new_val = sum(addends) if all(a is not None for a in addends) else default_value
                elif macro_text.lower() == 'race':
                    # source field may contain 'race' and 'ethnicity' source
                    # fields from which to derive final 'race' output value
                    source_field_names: list[str] = [s.strip() for s in next(csv.reader([source_field.strip(' []')]))]
                    if not source_field_names or len(source_field_names) > 2:
                        msg = (
                            f'Invalid source field "{source_field}" for "race" macro in source file ' +
                            f'{source_record["source_file_name"]}, must be single field OR comma-separated ' +
                            '(csv) string within square brackets specifying race and ethnicity fields such as ' +
                            '[race, ethnicity]'
                        )
                        _logger.critical(msg)
                        raise RuntimeError(msg)
                    source_race: str = source_record.get(source_field_names[0])
                    source_race = (
                        C3dcEtl.MULTIPLE_VALUE_DELIMITER.join(source_race)
                            if isinstance(source_race, (list, set, tuple))
                            else source_race
                    )
                    source_ethnicity: str = (
                        source_record.get(source_field_names[1]) if len(source_field_names) == 2 else ''
                    )
                    source_ethnicity = (
                        C3dcEtl.MULTIPLE_VALUE_DELIMITER.join(source_ethnicity)
                            if isinstance(source_ethnicity, (list, set, tuple))
                            else source_ethnicity
                    )
                    races: list[str] = self._get_race(source_race, source_ethnicity)
                    race: str
                    valid_races: list[str] = []
                    output_field: str = mapping.get('output_field')
                    for race in races:
                        case_matched_race: str = self._case_match_json_schema_enum_value(output_field, race)
                        if case_matched_race:
                            valid_races.append(case_matched_race)
                        else:
                            msg = (
                                f'Invalid source value "{race}" in "{source_field}" for "race" macro in source file ' +
                                f'{source_record["source_file_name"]}, not found in data dictionary'
                            )
                            _logger.warning(msg)
                    new_val = (
                        sorted(valid_races)
                            if valid_races and all(r not in ('', None) for r in valid_races)
                            else default_value
                    )
                new_vals[i] = new_val

            if new_value == '{find_enum_value}' and new_val is None:
                # enum lookup didn't find match, continue on to next replacement entry if available
                # ex: 0001/0 not in diagnosis.diagnosis enum value list but has replacement 8000/0
                # manual replacement ('Replacement Values' col in mapping) will follow the enum lookup
                # in the mapping's replacment_entries section
                continue

            new_value = new_vals if isinstance(new_value, (list, set, tuple)) else new_vals[0]
            if C3dcEtl.is_replacement_match(source_field, source_record, old_value):
                output_value = new_value
                break

            if isinstance(source_value, (list, set, tuple)):
                raise RuntimeError(f'Source value is collection: {source_record}')

        return output_value

    def _get_type_group_index_mappings(
        self,
        transformation: dict[str, any],
        node_type: C3dcEtlModelNode,
        clear_cache: bool = False
    ) -> dict[str, list[dict[str, any]]]:
        """
        Collate and return mappings of specified tranformation by type group index if defined, for example if
        multiple mappings of the same type are needed for a single transformation operation as may be the case
        for reference files, initial diagnosis + relapse diagnoses, etc
        """
        type_group_index_mappings: dict[str, list[dict[str, any]]] = transformation.get(
            '_type_group_index_mappings',
            {}
        ).get(node_type, {})
        if type_group_index_mappings and not clear_cache:
            return type_group_index_mappings

        # get mappings for specified node type
        type_group_index: str
        type_group_index_mappings: dict[str, list[dict[str, any]]] = {}
        mappings: list[dict[str, any]] = [
            m for m in transformation.get('mappings', []) if m.get('output_field', '').startswith(f'{node_type}.')
        ]
        mapping: dict[str, any]
        for mapping in mappings:
            type_group_indexes: list[str] = [i.strip() for i in str(mapping.get('type_group_index', '*')).split(',')]
            for type_group_index in type_group_indexes:
                type_group_index_mappings[type_group_index] = type_group_index_mappings.get(type_group_index) or []
                type_group_index_mappings[type_group_index].append(mapping)

        # replicate base/default mapping collection to remaining mapping groups, using
        # custom sort order to maintain order by type group index with exception for '*'
        type_group_index_mappings = dict(
            sorted(
                type_group_index_mappings.items(),
                key=lambda item: 0 if item[0] in ('', '*') else int(item[0])
            )
        )
        if not type_group_index_mappings:
            return type_group_index_mappings

        base_mappings: list[dict[str, any]] = type_group_index_mappings.get('*', [])
        non_base_mappings: dict[str, list[dict[str, any]]] = {
            k:v for k,v in type_group_index_mappings.items() if k != '*'
        }

        if base_mappings and non_base_mappings:
            for type_group_index, mappings in non_base_mappings.items():
                for mapping in reversed(base_mappings):
                    if not any(m for m in mappings if m.get('output_field') == mapping.get('output_field')):
                        mappings.insert(0, mapping)

        # the base/default mapping group is only needed if it's the only group so remove if there are multiple groups
        if non_base_mappings:
            type_group_index_mappings = non_base_mappings

        # cache the type group index mapping in the tranformation object for future re-use
        transformation['_type_group_index_mappings'] = transformation.get('_type_group_index_mappings', {})
        transformation['_type_group_index_mappings'][node_type] = type_group_index_mappings
        return type_group_index_mappings

    def _get_allowed_values(self, mapping: dict[str, any]) -> set[str]:
        """ Get allowed values for specified mapping """
        replacement_entries: list[dict[str, str]] = mapping.get('replacement_values', [])
        allowed_values: set[str] = set(
            r.get('old_value') for r in replacement_entries
                if r.get('old_value') not in ('+', '*', None) and r.get('new_value')
        )

        # don't include default value in allowed values unless output property is enum
        default_value: any = mapping.get('default_value')
        if default_value is not None and mapping.get('output_field') in self._json_schema_property_enum_values:
            if not isinstance(default_value, (list, set, tuple)):
                default_value = set([default_value])
            allowed_values.update(default_value)

        # check if all values are allowed if output property is enum
        if any(
            r.get('old_value') in ('+', '*') and r.get('new_value') == '{find_enum_value}' for r in replacement_entries
        ):
            # all enum values for mapped output field are allowed
            enum_code_values: dict[str, str] = self._json_schema_property_enum_code_values.get(
                mapping.get('output_field'),
                {}
            )
            allowed_values.update(set(enum_code_values.keys()))

        # empty string ("") and None are treated equally for matching/comparison purposes
        if '' in allowed_values:
            allowed_values.add(None)

        return allowed_values

    def _find_source_field(self, transformation: dict[str, any], output_field: str) -> str:
        """
        Find source field for specified output field in transformation mappings. If the number of matches found
        is not equal to 1 then None will be returned.
        """
        matched_mappings: list[dict[str, any]] = [
            m for m in transformation.get('mappings', []) if m.get('output_field', '').strip() == output_field
        ]
        return matched_mappings[0].get('source_field', '').strip() if len(matched_mappings) == 1 else None

    def _transform_record_default(
        self,
        transformation: dict[str, any],
        node_type: C3dcEtlModelNode,
        source_record: dict[str, any] = None
    ) -> list[dict[str, any]]:
        """ Transform and return result after applying non row-mapped transformation to specified source record """
        source_record = source_record or {}
        output_records: list[dict[str, any]] = []

        type_group_index_mappings: dict[str, list[dict[str, any]]] = self._get_type_group_index_mappings(
            transformation,
            node_type
        )
        if not type_group_index_mappings:
            _logger.warning('No mappings found for type %s, unable to transform record', node_type)
            return []

        # a single source field can be mapped to multiple target fields, for example
        # 'First Event' => survival.first_event, survival.event_free_survival_status
        # so pre-load all allowed values for each source field to determine when a
        # source record contains invalid/unmapped values and should be skipped/ignored
        allowed_values: set[str]
        source_field_allowed_values: dict[str, list[str]] = {}
        type_group_index: str
        mappings: list[dict[str, any]]
        mapping: dict[str, any]
        for type_group_index, mappings in type_group_index_mappings.items():
            for mapping in mappings:
                source_field: str = mapping.get('source_field')
                if source_field.startswith('[') and source_field.endswith(']'):
                    # don't constrain string substitution mappings like '[string_literal]'
                    continue
                source_field_allowed_values[source_field] = source_field_allowed_values.get(source_field, [])
                source_field_allowed_values[source_field].extend(self._get_allowed_values(mapping))

        output_source_field_map: dict[str, str] = {}

        # if there are multiple type group indexes (multiple records are mapped) then default
        # values to base ("*") values first and then overwrite individual mapped fields for
        # each mapping sub-group specified for that type group index
        base_record: dict[str, any] = {}
        for type_group_index, mappings in type_group_index_mappings.items():
            output_record: dict[str, any] = {}
            output_record.update(base_record)
            for mapping in mappings:
                output_field: str = mapping.get('output_field')
                output_field_property: str = output_field[len(f'{node_type}.'):]

                default_value: any = mapping.get('default_value')

                source_field: str = mapping.get('source_field')
                source_value: str = source_record.get(source_field, None)
                if source_value in ('', None) and default_value is not None:
                    source_value = default_value

                if output_field not in output_source_field_map:
                    output_source_field_map[output_field] = source_field

                # check source value against all of this source field's mapped replacement values
                # in case there are multiple output field mappings for this source field
                allowed_values = source_field_allowed_values.get(source_field, set())
                source_value_allowed: bool = C3dcEtl.is_allowed_value(source_value, allowed_values)
                if allowed_values and not source_value_allowed:
                    _logger.warning(
                        (
                            '"%s" not specified as allowed value (old_value) in transformation(s) for source field ' +
                            '"%s", source record "%s"'
                        ),
                        source_value if source_value is not None else '',
                        source_field,
                        source_record.get('source_file_name')
                    )
                    continue

                # check source value against mapped replacement values for this particular output field mapping
                allowed_values = self._get_allowed_values(mapping)
                source_value_allowed = C3dcEtl.is_allowed_value(source_value, allowed_values)
                if allowed_values and not source_value_allowed:
                    _logger.info(
                        'value "%s" not allowed for source field "%s"',
                        source_value if source_value is not None else '',
                        source_field
                    )
                    continue

                output_value: any = self._get_mapped_output_value(mapping, source_record)

                output_record[output_field_property] = output_value if output_value is not None else source_value

            # verify that record is valid and contains all required properties
            record_valid: bool = True
            required_properties: dict[str, any] = self._get_json_schema_node_required_properties(node_type)
            required_property: str
            for required_property in required_properties:
                schema_field: str = f'{node_type}.{required_property}'
                required_property_value: any = output_record.get(required_property, None)
                if (
                    required_property_value in ('', None, [])
                    or
                    isinstance(required_property_value, list) and all(v in ('', None) for v in required_property_value)
                ):
                    record_valid = False
                    _logger.warning(
                        'Required output field "%s" (source field "%s") is null/empty for source record file "%s"',
                        schema_field,
                        output_source_field_map.get(schema_field, '*not mapped*'),
                        source_record.get("source_file_name")
                    )

            if not record_valid:
                # record failed validation, move on to next type group index
                continue

            if output_record:
                output_records.append(output_record)
            if type_group_index == 0:
                base_record.update(output_record)

        return output_records

    def _transform_record_row_mapped(
        self,
        transformation: dict[str, any],
        node_type: C3dcEtlModelNode,
        source_record: dict[str, any]
    ) -> list[dict[str, any]]:
        """ get row-mapped records for specified transformation, node type and source_record """
        if node_type not in self._row_mapped_node_builders:
            raise RuntimeError(f'No row-mapped builder defined for node type "{node_type}"')
        builder: C3dcRowMappedBuilder = self._row_mapped_node_builders[node_type]
        builder.mappings = self._row_mapped_node_mappings[node_type][transformation['name']]
        return builder.get_records(source_record)

    def _transform_record_treatment(
        self,
        transformation: dict[str, any],
        source_record: dict[str, any]
    ) -> list[dict[str, any]]:
        """ get treatment records for specified source record """
        return self._transform_record_row_mapped(transformation, C3dcEtlModelNode.TREATMENT, source_record)

    def _transform_record_treatment_response(
        self,
        transformation: dict[str, any],
        source_record: dict[str, any]
    ) -> list[dict[str, any]]:
        """ get treatment response records for specified source record """
        return self._transform_record_row_mapped(transformation, C3dcEtlModelNode.TREATMENT_RESPONSE, source_record)

    def _build_node(
        self,
        transformation: dict[str, any],
        node_type: C3dcEtlModelNode,
        source_record: dict[str, any] = None
    ) -> list[dict[str, any]]:
        """
        Build and return specified C3DC model node type. If a custom method named '_transform_record_{node_type}'
        is found then that will be called, otherwise the default base method will be called for all node types
        """
        transform_method_name: str = f'_transform_record_{node_type}'
        transform_method: any = getattr(self, transform_method_name, lambda: None)
        if (
            transform_method is None or
            not hasattr(self, transform_method_name) or
            not callable(transform_method)
        ):
            return self._transform_record_default(transformation, node_type, source_record)

        return transform_method(transformation, source_record)

    def _build_sub_source_records(
        self,
        source_record: dict[str, any],
        node_props: dict[str, any]
    ) -> list[dict[str, any]]:
        """
        If source record contains enum fields having multiple values separated by delimiter then process as multiple
        source records for each distinct value with all other properties kept the same except record id, which must
        be unique for each newly created record
        """
        # check each property in the source record that is an enum/string (as opposed to enum/array)
        # where ';' isn't present in the list of permissible values. if the source value contains ';'
        # then create a sub source record for each distinct value otherwise return empty collection
        sub_source_records: list[dict[str, any]] = []
        sub_src_rec_enum_props: set[str] = {
            p for p in self._sub_source_record_enum_properties if p.startswith(f'{node_props["type"]}.')
        }
        sub_src_rec_enum_prop: str
        for sub_src_rec_enum_prop in sub_src_rec_enum_props:
            source_field: str = self._sub_source_record_enum_properties.get(sub_src_rec_enum_prop)
            if not source_field or C3dcEtl.MULTIPLE_VALUE_DELIMITER not in str(source_record.get(source_field, '')):
                continue
            sub_src_ids_vals: dict[str, str] = {
                f'{source_record[node_props["source_id_field"]]}_{k}':v for k,v in dict(
                    enumerate(
                        sorted(
                            set(
                                s.strip() for s in source_record[source_field].split(
                                    C3dcEtl.MULTIPLE_VALUE_DELIMITER
                                ) if s.strip()
                            )
                        ),
                        1
                    )
                ).items()
            }
            _logger.info(
                (
                    '"%s" "%s" has %d distinct delimited value(s) for "%s" ("%s"), creating separate record per value'
                ),
                node_props['type'],
                source_record[node_props["source_id_field"]],
                len(sub_src_ids_vals),
                sub_src_rec_enum_prop.partition('.')[2],
                source_field
            )
            sub_src_id: str
            sub_src_val: str
            for sub_src_id, sub_src_val in sub_src_ids_vals.items():
                src_rec_clone: dict[str, any] = json.loads(json.dumps(source_record))
                src_rec_clone[node_props['source_id_field']] = sub_src_id
                src_rec_clone[source_field] = sub_src_val
                sub_source_records.append(src_rec_clone)
        return sub_source_records

    def _transform_source_data(self, study_id: str, transformation: dict[str, any]) -> dict[str, any]:
        """ Transform and return ETL data transformed using rules specified in config """
        _logger.info('Transforming source data')
        if not self._raw_etl_data_objects.get(study_id, {}).get(transformation.get('name'), []):
            self._load_source_data(study_id, transformation)
            if not self._raw_etl_data_objects[study_id][transformation.get('name')]:
                raise RuntimeError(f'No data loaded to transform for study {study_id}')

        participant_id_field: str = f'{C3dcEtlModelNode.PARTICIPANT}.{C3dcEtlModelNode.PARTICIPANT}_id'
        subject_id_field: str = self._find_source_field(transformation, participant_id_field)
        if not subject_id_field:
            raise RuntimeError(
                f'Unable to find single source mapping for "{participant_id_field}" in transformation mappings; ' +
                f'"{participant_id_field}" is either not mapped or is mapped multple times'
            )

        # define row-mapped nodes like treatment and treatment response separately for logging purposes
        row_mapped_nodes: tuple[C3dcEtlModelNode, ...] = (
            C3dcEtlModelNode.TREATMENT,
            C3dcEtlModelNode.TREATMENT_RESPONSE
        )
        row_mapped_node: C3dcEtlModelNode

        nodes: dict[C3dcEtlModelNode, dict[str, any]] = {
            C3dcEtlModelNode.DIAGNOSIS: {},
            C3dcEtlModelNode.PARTICIPANT: {},
            C3dcEtlModelNode.REFERENCE_FILE: {},
            C3dcEtlModelNode.STUDY: {},
            C3dcEtlModelNode.SURVIVAL: {}
        }
        # append row-mapped nodes to main nodes collection
        for row_mapped_node in row_mapped_nodes:
            nodes[row_mapped_node] = {}

        node: C3dcEtlModelNode
        node_props: dict[str, any]
        for node, node_props in nodes.items():
            node_props['harmonized_records'] = []
            node_props['type'] = node
            node_props['id_field'] = f'{node}_id'
            node_props['id_field_full'] = f'{node}.{node_props["id_field"]}'
            node_props['source_id_field'] = subject_id_field

        # build study node and add to node collection
        study: dict[str, any] = self._build_node(transformation, C3dcEtlModelNode.STUDY)
        if len(study) != 1:
            raise RuntimeError(f'Unexpected number of study nodes built ({len(study)}), check mapping')
        study = study[0]
        study[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field_full']] = []
        study[nodes[C3dcEtlModelNode.REFERENCE_FILE]['id_field_full']] = []

        # verify mappings defined for row-mapped nodes
        for row_mapped_node in row_mapped_nodes:
            mappings: list[dict[str, any]] = self._row_mapped_node_mappings.get(row_mapped_node, {}).get(
                transformation['name']
            )
            if not mappings:
                _logger.warning(
                    '"%s" builder mappings not specified, "%s" records will not be harmonized',
                    row_mapped_node,
                    row_mapped_node
                )

        # add observation and participant records to match source data records
        rec: dict[str, any]
        raw_processed: int = 0
        for rec in self._raw_etl_data_objects[study_id][transformation.get('name')]:
            raw_processed += 1
            if raw_processed % 100 == 0:
                _logger.info('%d raw records processed', raw_processed)
            participant: dict[str, any]
            participants: list[dict[str, any]] = self._build_node(transformation, C3dcEtlModelNode.PARTICIPANT, rec)
            if len(participants) != 1:
                _logger.warning(
                    '%s (%s): Unexpected number of participant nodes (%d) built for sourced record %d, excluding',
                    transformation.get('name'),
                    study_id,
                    len(participants),
                    rec['source_file_name']
                )
                participant = None
                continue
            participant = participants[0]

            node_observations: list[C3dcEtlModelNode] = [
                C3dcEtlModelNode.DIAGNOSIS,
                C3dcEtlModelNode.SURVIVAL,
                C3dcEtlModelNode.TREATMENT,
                C3dcEtlModelNode.TREATMENT_RESPONSE
            ]
            for node in node_observations:
                # make sure relationship collection is defined, even if no records are added
                participant[nodes[node]['id_field_full']] = []

                sub_src_rec: dict[str, any]
                for sub_src_rec in self._build_sub_source_records(rec, nodes[node]) or [rec]:
                    harmonized_recs: list[dict[str, any]] = self._build_node(transformation, node, sub_src_rec)
                    if (
                        not harmonized_recs
                        and
                        node not in (C3dcEtlModelNode.TREATMENT, C3dcEtlModelNode.TREATMENT_RESPONSE)
                    ):
                        _logger.warning(
                            '%s (%s): Unable to build "%s" node for source record "%s"',
                            transformation.get('name'),
                            study_id,
                            node,
                            sub_src_rec['source_file_name']
                        )

                    harmonized_rec: dict[str, any]
                    for harmonized_rec in harmonized_recs:
                        # set referential ids for this observation node and parent participant
                        harmonized_rec[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field_full']] = participant[
                            nodes[C3dcEtlModelNode.PARTICIPANT]['id_field']
                        ]
                        participant[nodes[node]['id_field_full']].append(harmonized_rec[nodes[node]['id_field']])

                    # populate final transformed record collection for this observation node
                    nodes[node]['harmonized_records'].extend(harmonized_recs)

            participant[nodes[C3dcEtlModelNode.STUDY]['id_field_full']] = study[
                nodes[C3dcEtlModelNode.STUDY]['id_field']
            ]
            study[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field_full']].append(
                participant[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field']]
            )
            nodes[C3dcEtlModelNode.PARTICIPANT]['harmonized_records'].append(participant)

        # get reference file mappings for harmonized participants then add to base transformation mappings
        self._append_reference_file_fieldset_mappings(
            study_id,
            transformation,
            nodes[C3dcEtlModelNode.PARTICIPANT]['harmonized_records']
        )

        # build reference file nodes and add to node collection
        reference_files: list[dict[str, any]] = self._build_node(transformation, C3dcEtlModelNode.REFERENCE_FILE)
        reference_file: dict[str, any]
        for reference_file in reference_files:
            reference_file[nodes[C3dcEtlModelNode.STUDY]['id_field_full']] = study[
                nodes[C3dcEtlModelNode.STUDY]['id_field']
            ]
            study[nodes[C3dcEtlModelNode.REFERENCE_FILE]['id_field_full']].append(
                reference_file[nodes[C3dcEtlModelNode.REFERENCE_FILE]['id_field']]
            )
        nodes[C3dcEtlModelNode.REFERENCE_FILE]['harmonized_records'].extend(reference_files)

        # log record counts for row-mapped nodes
        for row_mapped_node in row_mapped_nodes:
            _logger.info(
                '%d "%s" record(s) harmonized for %d distinct subject(s)',
                len(nodes[row_mapped_node]['harmonized_records']),
                row_mapped_node,
                len(
                    set(
                        r['participant.participant_id']
                            for r in nodes[row_mapped_node]['harmonized_records']
                    )
                )
            )

        # check for participants having treatment responses but not treatments
        participants_w_response_wo_treatment: list[dict[str, any]] = [
            p for p in nodes[C3dcEtlModelNode.PARTICIPANT]['harmonized_records']
                if (
                    p[nodes[C3dcEtlModelNode.TREATMENT_RESPONSE]['id_field_full']]
                    and
                    not p[nodes[C3dcEtlModelNode.TREATMENT]['id_field_full']]
                )
        ]
        if participants_w_response_wo_treatment:
            _logger.warning(
                '%d participants assigned "%s" records without "%s" records:',
                len(participants_w_response_wo_treatment),
                C3dcEtlModelNode.TREATMENT_RESPONSE,
                C3dcEtlModelNode.TREATMENT
            )
            _logger.warning(
                [p[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field']] for p in participants_w_response_wo_treatment]
            )
            remission_response_participant_ids_all: list[str] = [
                r[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field_full']]
                    for r in nodes[C3dcEtlModelNode.TREATMENT_RESPONSE]['harmonized_records']
                        if r['response'] == 'Complete Remission'
            ]
            remission_participant_ids: list[str] = [
                pid for pid in remission_response_participant_ids_all
                    if pid in [
                        p[nodes[C3dcEtlModelNode.PARTICIPANT]['id_field']] for p in participants_w_response_wo_treatment
                    ]
            ]
            _logger.warning(
                '%d participants having "response" set to "Complete Remission":',
                len(remission_participant_ids)
            )
            _logger.warning(remission_participant_ids)

        # check for dupe ids
        for node, node_props in nodes.items():
            node_rec: dict[str, any]
            id_cache: set[str] = set()
            dupe_ids: set[str] = set()
            for node_rec in node_props['harmonized_records']:
                if node_rec[node_props['id_field']] in id_cache:
                    dupe_ids.add(node_rec[node_props['id_field']])
                id_cache.add(node_rec[node_props['id_field']])
            if dupe_ids:
                raise RuntimeError(f'Duplicate {node} id(s) found: {dupe_ids}')

        # attach the main study object
        nodes[C3dcEtlModelNode.STUDY]['harmonized_records'].append(study)

        self._json_etl_data_sets[study_id] = self._json_etl_data_sets.get(study_id) or {}
        self._json_etl_data_sets[study_id][transformation.get('name')] = {
            C3dcEtlModelNode.get_pluralized_node_name(k):v['harmonized_records'] for k,v in nodes.items()
        }

        _logger.info(
            '%s records built for transformation "%s"',
            ', '.join(f'{len(v["harmonized_records"])} {k}' for k,v in nodes.items()),
            transformation.get("name")
        )

        return self._json_etl_data_sets[study_id][transformation.get('name')]

    def _create_json_etl_file(self, study_id: str, transformation: dict[str, any]) -> None:
        """ Create JSON ETL data file for specified raw source data set """
        _logger.info('Creating JSON ETL data file for transformation %s (%s)', transformation.get('name'), study_id)
        self._random = random.Random()
        self._random.seed(transformation.get('uuid_seed', None))
        self._load_source_data(study_id, transformation)
        self._transform_source_data(study_id, transformation)
        self._save_json_etl_data(study_id, transformation)


def print_usage() -> None:
    """ Print script usage """
    _logger.info('usage: python %s [optional config file name/path if not .env]', sys.argv[0])


def main() -> None:
    """ Script entry point """
    if len(sys.argv) > 2:
        print_usage()
        return

    c3dc_file_manager: C3dcFileManager = C3dcFileManager()
    config_file: str = sys.argv[1] if len(sys.argv) == 2 else '.env'
    if not c3dc_file_manager.file_exists(config_file):
        raise FileNotFoundError(f'Config file "{config_file}" not found')
    config: dict[str, str] = dotenv.dotenv_values(
        stream=io.StringIO(c3dc_file_manager.read_file(config_file).decode('utf-8'))
    )
    etl: C3dcEtl = C3dcEtl(config)
    etl.create_json_etl_files()
    etl.validate_json_etl_data()


if __name__ == '__main__':
    main()
