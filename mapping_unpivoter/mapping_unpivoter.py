""" Unpivoter for to convert internal field mapping to deliverable transformation mapping """
from __future__ import annotations

import csv
import hashlib
import json
import logging
import logging.config
import os
import pathlib
import sys
from urllib.parse import urlparse, ParseResult
from urllib.request import url2pathname

from contextlib import ContextDecorator
import dotenv
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import requests


_logger = logging.getLogger(__name__)
if _logger.hasHandlers():
    _logger.handlers.clear()

logging.config.dictConfig({
    "version": 1,
    "disable_existing_loggers": True,
    "formatters": {
        "standard": {
            "format": "%(asctime)s [%(levelname)s]: %(message)s"
        }
    },
    "handlers": {
        "console": {
            "level": "DEBUG",
            "formatter": "standard",
            "class": "logging.StreamHandler",
            "stream": "ext://sys.stdout",  # Default is stderr
        },
        "file": {
            "level": "DEBUG",
            "formatter": "standard",
            "class": "logging.FileHandler",
            "filename": "mapping_unpivoter.log",
            "mode": "w"
        }
    },
    "loggers": {
        "": { # root logger
            "handlers": ["console", "file"],
            "level": "DEBUG",
            "propagate": False
        },
        "__main__": {
            "handlers": ["console", "file"],
            "level": "DEBUG",
            "propagate": False
        }
    }
})


class MappingUnpivoter(ContextDecorator):
    """
    Unpivot tabular transformation field mappings to JSON format. Use as context manager e.g.:

    config: dict[str, str] = dotenv.dotenv_values('/path/to/.env')
    mapping_unpivoter: MappingUnpivoter
    with MappingUnpivoter(config) as mapping_unpivoter:
        mapping_unpivoter.unpivot_transformation_mappings()
    """
    TRUE_STRINGS: tuple[str, ...] = (str(True).lower(), 't', 'yes', 'y', '1')

    REF_FILE_CATEGORY_PROG_SOURCE_CODE: str = 'programmatic source code'
    REF_FILE_CATEGORY_XFORM_MAP: str = 'transformation/mapping'
    REF_FILE_CATEGORY_OUTPUT_SCHEMA: str = 'output schema'
    REF_FILE_CATEGORY_INPUT_SOURCE_DATA: str = 'input source data'

    def __init__(self, config: dict[str, str]) -> None:
        self._config: dict[str, str] = config
        self._version: str = config.get('VERSION')
        self._json_schema_url: str = config.get('JSON_SCHEMA_URL')
        self._json_schema: dict[str, any] = {}
        self._node_type_properties: list[str] = []
        self._output_file: str = config.get('OUTPUT_FILE')
        self._etl_script_file: str = config.get('ETL_SCRIPT_FILE')
        self._json_schema_file = './schema.json'
        self._transformation_mappings_files: list[dict[str, any]] = json.loads(
            config.get('TRANSFORMATION_MAPPINGS_FILES', '[]')
        )
        self._transformation_config: dict[str, any] = {'version': self._version, 'transformations': []}

    def __enter__(self) -> None:
        """ download the JSON schema to a local file """
        self._load_json_schema()
        self._cache_node_type_properties()
        return self

    def __exit__(self, exc_type, exc, exc_tb) -> None:
        """ delete the local copy of the JSON schema"""
        if os.path.exists(self._json_schema_file):
            os.remove(self._json_schema_file)

    @staticmethod
    def is_number(value: str):
        """ Determine whether specified string is number (float or int) """
        try:
            float(value)
        except (TypeError, ValueError):
            return False
        return True

    @staticmethod
    def url_to_path(url: str) -> str:
        """ Convert specified URL to path specific to local platform """
        url_parts: ParseResult = urlparse(url)
        host = f"{os.path.sep}{os.path.sep}{url_parts.netloc}{os.path.sep}"
        return os.path.normpath(os.path.join(host, url2pathname(url_parts.path)))

    @staticmethod
    def get_url_content(url: str, local_save_path: str = None) -> any:
        """ Retrieve and return contents from specified URL """
        url_content: str | bytes | bytearray
        if url.startswith('file://'):
            with open(MappingUnpivoter.url_to_path(url), 'rb') as local_file:
                url_content = local_file.read()
        else:
            with requests.get(url, timeout=30) as response:
                response.raise_for_status()
                url_content = response.content

        if local_save_path:
            with open(local_save_path, 'wb') as write_fp:
                write_fp.write(url_content)

        return url_content

    @property
    def transformation_config(self) -> dict[str, any]:
        """
        Get internal schema object, building if needed
        """
        return self._transformation_config

    def load_transformation_config_output_file(self) -> None:
        """ Load existing transformation config from (output) file specified in config """
        _logger.info('Loading transformation config from "%s"', self._output_file)

        if not os.path.isfile(self._output_file):
            raise RuntimeError(f'Transformation/mapping output file "{self._output_file}" not found')

        with open(self._output_file, 'r', encoding='utf-8') as fp:
            self._transformation_config = json.load(fp)

    def save_transformation_config_output_file(self) -> None:
        """ Save transformation config to output file specified in config """
        _logger.info('Saving transformation config to "%s"', self._output_file)
        with open(self._output_file, mode='w', encoding='utf-8') as fp:
            json.dump(self._transformation_config, fp, indent=4)

    def get_transformation_mappings_file_records(
        self,
        transformation_mappings_file: dict[str, any]
    ) -> list[dict[str, any]]:
        """ Return contents of transformation mappings file as list of dicts """
        records: list[dict[str, any]] = []
        if pathlib.Path(transformation_mappings_file['mappings_file']).suffix.lower() == '.csv':
            with open(transformation_mappings_file['mappings_file'], encoding='utf-8') as csv_fp:
                reader: csv.DictReader = csv.DictReader(csv_fp)
                record: dict[str, any]
                for record in reader:
                    records.append(record)
        elif pathlib.Path(transformation_mappings_file['mappings_file']).suffix.lower() == '.xlsx':
            sheet_name: str = transformation_mappings_file.get('mappings_file_sheet')
            sheet_name = sheet_name[:31]
            wb: Workbook = openpyxl.load_workbook(transformation_mappings_file['mappings_file'], data_only=True)
            if sheet_name not in wb.sheetnames:
                raise RuntimeError(f'Worksheet "{sheet_name}" not found in workbook worksheet list: {wb.sheetnames}')
            ws: Worksheet = wb[sheet_name] if sheet_name else wb[0]
            row: str
            cols: list[str] = []
            for row in ws.iter_rows():
                if not cols:
                    cols = [cell.value for cell in row]
                    continue
                record: dict[str, any] = {}
                col_num: int
                for col_num, cell in enumerate(row):
                    value: any = cell.value
                    if MappingUnpivoter.is_number(value):
                        value = int(float(value)) if float(value).is_integer() else float(value)
                    record[cols[col_num]] = str(value) if value is not None else ''
                records.append(record)
        else:
            raise RuntimeError(
                f'Unsupported transformation mappings file type: "{transformation_mappings_file["mappings_file"]}"'
            )
        return records

    def unpivot_transformation_mappings(self) -> None:
        """ Unpivot tablular mapping data to JSON """
        _logger.info('Unpivoting transformation mappings')
        errors: list[str] = []
        err_msg: str
        self._transformation_config['transformations'] = self._transformation_config.get('transformation', [])
        self._transformation_config['transformations'].clear()
        transformation_mappings_file: dict[str, any]
        for transformation_mappings_file in self._transformation_mappings_files:
            _logger.info('Loading transformation mappings file %s', transformation_mappings_file['mappings_file'])
            records: list[dict[str, any]] = self.get_transformation_mappings_file_records(transformation_mappings_file)
            record: dict[str, any]
            unpivoted_mappings: list[dict[str, any]] = []
            for record in records:
                # unpivot pivoted mapping
                unpivoted_mapping: dict[str, any] = self._unpivot_mapping(record)
                if not unpivoted_mapping:
                    continue

                # check for consistent default values for this output field's mappings
                if unpivoted_mapping['default_value'] is not None:
                    existing_default_values: set[any] = set()
                    default_value: any
                    for default_value in [
                        m['default_value'] for m in unpivoted_mappings
                            if m['output_field'] == unpivoted_mapping['output_field'] and
                                m['default_value'] is not None
                    ]:
                        if not isinstance(default_value, (list, set, tuple)):
                            default_value = set([default_value])
                        existing_default_values.update(default_value)
                    if isinstance(unpivoted_mapping['default_value'], (list, set, tuple)):
                        existing_default_values.update(unpivoted_mapping['default_value'])
                    else:
                        existing_default_values.add(unpivoted_mapping['default_value'])
                    if len(existing_default_values) > 1:
                        raise RuntimeError(
                            f'Invalid mapping for output field "{unpivoted_mapping["output_field"]}", ' +
                            f'multiple default values specified: {existing_default_values}')

                # check for existing unpivoted mapping for this pivoted mapping's output field and type
                # group, if found then this field mapping has multiple replacement values so append
                existing_unpivoted_mappings: list[dict[str, any]] = [
                    m for m in unpivoted_mappings
                        if m['output_field'] == unpivoted_mapping['output_field'] and
                            m['type_group_index'] == unpivoted_mapping['type_group_index']
                ]
                if not existing_unpivoted_mappings:
                    unpivoted_mappings.append(unpivoted_mapping)
                else:
                    source_field_count: int = len(set(m.get('source_field') for m in existing_unpivoted_mappings))
                    if source_field_count != 1:
                        err_msg = (
                            f'Invalid mapping: unexpected number of source fields ({source_field_count}) ' +
                            f'for output field {unpivoted_mapping["output_field"]}'
                        )
                        errors.append(err_msg)
                        continue

                    existing_unpivoted_mappings[0].get('replacement_values').extend(
                        unpivoted_mapping['replacement_values']
                    )

            transformation: dict[str, any] = {
                'name': transformation_mappings_file['transformation_name'],
                'mappings': unpivoted_mappings
            }
            self._transformation_config['transformations'].append(transformation)

        # save transformation config to output file specified in config
        _logger.info('Saving unpivoted transformation mappings to %s', self._output_file)
        self.save_transformation_config_output_file()

    def update_reference_file_mappings(self) -> None:
        """
        Update reference file size and md5 sum mappings for all reference files specified
        in configuration: ETL script, schema, transformation/mapping, input source data
        """
        ref_files_to_update: dict[str, str] = {}
        ref_files_to_update[self._json_schema_file] = MappingUnpivoter.REF_FILE_CATEGORY_OUTPUT_SCHEMA
        if self._etl_script_file:
            ref_files_to_update[self._etl_script_file] = MappingUnpivoter.REF_FILE_CATEGORY_PROG_SOURCE_CODE
        for transformation_mappings_file in self._transformation_mappings_files:
            if transformation_mappings_file.get('source_data_file'):
                ref_files_to_update[transformation_mappings_file['source_data_file']] = \
                    MappingUnpivoter.REF_FILE_CATEGORY_INPUT_SOURCE_DATA

        ref_file_to_update: str
        ref_file_category: str
        for ref_file_to_update, ref_file_category in ref_files_to_update.items():
            self._update_reference_file_mappings(ref_file_to_update, ref_file_category)
        self.save_transformation_config_output_file()

        # The size and md5 hass of the transformation/mapping file will be calculated based on the properties for
        # the other ref files already being populated and the self-referential file size and md5 hash properties
        # for the transformation/mapping being set to 0 and '' respectively. This will allow deterministic
        # confirmation if needed in the future by calculating these properties for the file when the values are
        # set back to 0 and ''.
        self.load_transformation_config_output_file()
        self._update_reference_file_mappings(self._output_file, MappingUnpivoter.REF_FILE_CATEGORY_XFORM_MAP)
        self.save_transformation_config_output_file()

    def _update_reference_file_mappings(self, ref_file_path: str, ref_file_category: str) -> None:
        """
        Update mapping entries for reference file size and md5 sum for specified reference file and category.
        Return True if update performed successfully, False otherwise
        """
        _logger.info('Updating file size and md5sum for "%s" reference file "%s"', ref_file_category, ref_file_path)
        # get size and md5 hash of output file to set matching reference file properties
        ref_file_size: int = os.path.getsize(ref_file_path)

        md5: hashlib._Hash = hashlib.md5()
        with open(ref_file_path, mode='rb') as ref_file_fp:
            chunk: any
            for chunk in iter(lambda: ref_file_fp.read(4096), b''):
                md5.update(chunk)
        ref_file_md5sum: str = md5.hexdigest()

        ref_file_field_value_map: dict[str, any] = {
            'reference_file.file_size': ref_file_size,
            'reference_file.md5sum': ref_file_md5sum
        }

        ref_file_basename: str = os.path.basename(ref_file_path)

        # update file size and md5 values of matching transformation/mapping ref file entries for each transformation
        update_successful: bool = True
        for xform_cfg in self._transformation_config.get('transformations', []):
            xform_updated: bool = True
            reference_file_tgis: list[str] = []
            mapping: dict[str, any]
            # get type group index values of relevant transformation/mapping files
            # by matching reference file's file name property to our output file
            for mapping in [
                m for m in xform_cfg['mappings']
                    if m.get('type_group_index') and
                        m.get('output_field') == 'reference_file.file_name' and
                        any(r for r in m.get('replacement_values', []) if r.get('new_value') == ref_file_basename)
            ]:
                reference_file_tgis.append(mapping.get('type_group_index'))

            if not reference_file_tgis:
                _logger.warning(
                    'No reference file mapping for "%s" found in transformation "%s"',
                    ref_file_basename,
                    xform_cfg.get('name')
                )

            # set file size and md5sum values for each reference file set circumscribed by type group index
            tgi: str
            for tgi in set(reference_file_tgis):
                ref_file_size_md5sum_mappings: list[dict[str, any]] = self._get_ref_file_size_and_md5sum_mappings(
                    ref_file_path,
                    ref_file_category,
                    xform_cfg['mappings'],
                    tgi
                )
                if not ref_file_size_md5sum_mappings:
                    _logger.warning('Ref file size and md5sum mappings not found')
                output_field: str
                new_value: any
                values_updated: int = 0
                for output_field, new_value in ref_file_field_value_map.items():
                    ref_file_mapping: dict[str, any] = [
                        m for m in ref_file_size_md5sum_mappings if m.get('output_field') == output_field
                    ]
                    if not ref_file_mapping:
                        raise RuntimeError(
                            f'Mapping for "{output_field}" not found for transformation "{xform_cfg.get("name")}" ' +
                            f'(type group index {tgi})'
                        )
                    ref_file_mapping = ref_file_mapping[0]
                    replacement_value: dict[str, any]
                    replacement_value_updated: bool = False
                    for replacement_value in ref_file_mapping.get('replacement_values', []):
                        if replacement_value['new_value'] not in ('', 0):
                            raise RuntimeError(
                                f'Unexpected value for "new_value" in mapping replacement_values: {ref_file_mapping}'
                            )
                        _logger.info(
                            (
                                '\tSetting "%s" replacement "new_value" properties to "%s" for ' +
                                    'type_group_index "%s" in transformation "%s"'
                            ),
                            output_field,
                            new_value,
                            tgi,
                            xform_cfg.get('name')
                        )
                        replacement_value['new_value'] = new_value
                        replacement_value_updated = True
                    values_updated += (1 if replacement_value_updated else 0)
                if values_updated < len(ref_file_field_value_map):
                    xform_updated = False
            update_successful = update_successful and xform_updated
        if not update_successful:
            _logger.warning('Update of transformation/mapping reference file size and md5sum failed')

    def _load_json_schema(self) -> dict[str, any]:
        """ Load JSON schema from location specified in config """
        _logger.info('Loading JSON schema from %s', self._json_schema_url)
        # save local copy of json schema file to calculate size and md5 hash to update ref file mapping if needed
        self._json_schema = json.loads(MappingUnpivoter.get_url_content(self._json_schema_url, self._json_schema_file))
        return self._json_schema

    def _cache_node_type_properties(self) -> list[str]:
        """ Get list of node type properties in 'node_type.property_name' format and cache in instance var """
        _logger.info('Caching node type properties')

        if not self._json_schema:
            self._load_json_schema()
            if not self._json_schema:
                raise RuntimeError('Unable to cache node type properties; failed to load JSON schema')

        self._node_type_properties = []
        defs: dict[str, any] = self._json_schema.get('$defs', {})
        etl_node_name: str
        etl_node_obj: dict[str, any]
        # enumerate nodes in $defs
        for etl_node_name, etl_node_obj in defs.items():
            property_name: str
            # enumerate properties in 'properties'
            for property_name in etl_node_obj.get('properties', {}):
                self._node_type_properties.append(f'{etl_node_name}.{property_name}')
        if self._node_type_properties:
            _logger.info('Cached %d properties for %d node types', len(self._node_type_properties), len(defs.keys()))
        else:
            raise RuntimeError('No node type properties found to be cached')
        return self._node_type_properties

    def _unpivot_mapping(self, pivoted_mapping: dict[str, any]) -> dict[str, any]:
        """ Convert specified flat mapping loaded from tabular data file to JSON mapping """
        # if output field isn't set/defined then this mapping entry is for reference only and can be skipped
        target_variable_name: str = pivoted_mapping.get('Target Variable Name', '')
        if target_variable_name not in self._node_type_properties:
            _logger.warning(
                'Mapping has invalid "Target Variable Name", skipping: "%s"',
                pivoted_mapping.get('Target Variable Name', '')
            )

            return {}

        unpivoted_mapping: dict[str, any] = {}
        try:
            unpivoted_mapping['output_field'] = pivoted_mapping['Target Variable Name']
            unpivoted_mapping['source_field'] = pivoted_mapping['Source Variable Name']
            unpivoted_mapping['type_group_index'] = str(pivoted_mapping['Type Group Index'])
            unpivoted_mapping['default_value'] = json.loads(pivoted_mapping['Default Value If Null/Blank'] or 'null')
            unpivoted_mapping['replacement_values'] = unpivoted_mapping.get('replacement_values', [])
            unpivoted_mapping['replacement_values'].append(
                {
                    'old_value': json.loads(pivoted_mapping['Source Permissible Values Term'] or '""'),
                    'new_value': json.loads(pivoted_mapping['Target Permissible Values Term'] or '""')
                }
            )

            repl_vals: dict[str, str] = json.loads(pivoted_mapping.get('Replacement Values', '{}') or '{}')
            replacement_values: list[dict[str, str]] = [{'old_value': k, 'new_value': v} for k,v in repl_vals.items()]
            unpivoted_mapping['replacement_values'].extend(replacement_values)
        except json.decoder.JSONDecodeError as err:
            _logger.error('Error loading replacement values while unpivoting mapping:')
            _logger.error(err)
            _logger.error(pivoted_mapping)
            raise
        except TypeError as terr:
            _logger.error(terr)
            _logger.error(pivoted_mapping)
            raise
        return unpivoted_mapping

    def _get_ref_file_size_and_md5sum_mappings(
        self,
        ref_file: str,
        ref_file_category: str,
        mappings: list[dict[str, any]],
        type_group_index: str,
    ) -> list[dict[str, any]]:
        """ Get all 'reference_file.*' mappings in list of mappings for specified type group index value """
        ref_file_mappings: list[dict[str, any]] = [
            m for m in mappings
                if m.get('type_group_index') == type_group_index and
                    m.get('output_field', '').startswith('reference_file.')
        ]
        file_name_match: list[dict[str, any]] = [
            m for m in ref_file_mappings
                if m.get('output_field') == 'reference_file.file_name' and any(
                    r for r in m.get('replacement_values', [])
                        if r.get('new_value') == os.path.basename(ref_file)
                )
        ]
        file_cat_match: list[dict[str, any]] = [
            m for m in ref_file_mappings
                if m.get('output_field') == 'reference_file.file_category' and any(
                    r for r in m.get('replacement_values', [])
                        if r.get('new_value') == ref_file_category
                )
        ]
        file_size_match: list[dict[str, any]] = [
            m for m in ref_file_mappings
                if m.get('output_field') == 'reference_file.file_size' and any(
                    r for r in m.get('replacement_values', [])
                        if r.get('new_value') == 0
                )
        ]
        file_md5sum_match: list[dict[str, any]] = [
            m for m in ref_file_mappings
                if m.get('output_field') == 'reference_file.md5sum' and any(
                    r for r in m.get('replacement_values', [])
                        if r.get('new_value') == ''
                )
        ]
        if file_name_match and file_cat_match and file_size_match and file_md5sum_match:
            return [
                m for m in ref_file_mappings
                    if m.get('output_field') in ('reference_file.file_size', 'reference_file.md5sum')
            ]
        return []


VALID_METHODS: tuple[str, ...] = (
    MappingUnpivoter.unpivot_transformation_mappings.__name__,
    MappingUnpivoter.update_reference_file_mappings.__name__
)


def print_usage() -> None:
    """ Print script usage """
    _logger.info('usage: python %s [%s] [optional config file name/path]', sys.argv[0], '|'.join(VALID_METHODS))


def main() -> None:
    """ Script entry point """
    if len(sys.argv) not in (2, 3) or sys.argv[1] not in VALID_METHODS:
        print_usage()
        return

    config_file: str = sys.argv[2] if len(sys.argv) == 3 else '.env_mapping_unpivoter'
    if not os.path.exists(config_file):
        raise FileNotFoundError(f'Config file "{config_file}" not found')
    config: dict[str, str] = dotenv.dotenv_values(config_file)
    mapping_unpivoter: MappingUnpivoter
    with MappingUnpivoter(config) as mapping_unpivoter:
        if sys.argv[1] == MappingUnpivoter.unpivot_transformation_mappings.__name__:
            mapping_unpivoter.unpivot_transformation_mappings()
            if config.get('AUTO_UPDATE_REFERENCE_FILE_MAPPINGS', 'False').lower() in MappingUnpivoter.TRUE_STRINGS:
                mapping_unpivoter.update_reference_file_mappings()
                mapping_unpivoter.save_transformation_config_output_file()
        elif sys.argv[1] == MappingUnpivoter.update_reference_file_mappings.__name__:
            mapping_unpivoter.load_transformation_config_output_file()
            mapping_unpivoter.update_reference_file_mappings()
            mapping_unpivoter.save_transformation_config_output_file()


if __name__ == '__main__':
    main()
